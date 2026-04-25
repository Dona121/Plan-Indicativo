"""
Dashboard — Plan Indicativo 2024-2027
Seguimiento de ejecución física y financiera del Plan de Desarrollo.
"""

import io
import streamlit as st
import polars as pl
import polars.selectors as cs
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
import requests

# =========================================================================
# Paleta corporativa
# =========================================================================
COLORS = {
    "green_light":  "#17743d",
    "green_dark":   "#005931",
    "cyan":         "#47b1d5",
    "blue":         "#1754ab",
    "blue_dark":    "#003d6c",
    "orange":       "#d88c16",
    "orange_deep":  "#cf7000",
    "amber":        "#d37e00",
    "brown":        "#9b5b1e",
    "coral":        "#e68878",
}

# Fuentes oficiales:
#   Alkaline  -> fuente de display / titulares (carácter institucional)
#   Montserrat -> encabezados y UI
#   Open Sans  -> cuerpo, datos, formularios
FONT_DISPLAY = "Alkaline"
FONT_HEADING = "Montserrat"
FONT_BODY    = "Open Sans"
FONT_MONO    = "JetBrains Mono"  # apoyo para metadatos pequeños

# =========================================================================
# Configuración de página
# =========================================================================
st.set_page_config(
    page_title="Plan Indicativo 2024-2027",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================================================================
# CSS institucional
# =========================================================================
CSS = f"""
<style>
/* Montserrat y Open Sans oficiales desde Google Fonts.
   Alkaline no está disponible en Google Fonts, se carga desde un CDN libre;
   si falla, el stack hace fallback a Montserrat + serif. */
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600;700;800&family=Open+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

@font-face {{
    font-family: 'Alkaline';
    src: url('https://cdn.jsdelivr.net/gh/gogolapse/fonts@main/Alkaline/Alkaline.otf') format('opentype');
    font-weight: 400;
    font-style: normal;
    font-display: swap;
}}

:root {{
    --green-light: {COLORS["green_light"]};
    --green-dark:  {COLORS["green_dark"]};
    --cyan:        {COLORS["cyan"]};
    --blue:        {COLORS["blue"]};
    --blue-dark:   {COLORS["blue_dark"]};
    --orange:      {COLORS["orange"]};
    --orange-deep: {COLORS["orange_deep"]};
    --amber:       {COLORS["amber"]};
    --brown:       {COLORS["brown"]};
    --coral:       {COLORS["coral"]};

    --paper:       #ffffff;
    --ink:         #0d1b2a;
    --ink-mute:    #4a5a6a;
    --hairline:    #e3e3e1;
    --chip-bg:     #f1f1ef;
}}

/* Base tipográfica */
html, body, [class*="css"], .stApp {{
    font-family: '{FONT_BODY}', system-ui, sans-serif !important;
    color: var(--ink);
}}

.stApp {{
    background:
        radial-gradient(1200px 600px at 90% -10%, rgba(23,84,171,0.06), transparent 60%),
        radial-gradient(900px 500px at -10% 110%, rgba(216,140,22,0.05), transparent 60%),
        var(--paper);
}}

/* Encabezados H1..H6 con Montserrat (institucional, geométrica) */
h1, h2, h3, h4, h5, h6 {{
    font-family: '{FONT_HEADING}', Helvetica, sans-serif !important;
    color: var(--ink);
    letter-spacing: -0.01em;
}}
h1 {{ font-weight: 800 !important; }}
h2, h3 {{ font-weight: 700 !important; }}
h4, h5, h6 {{ font-weight: 600 !important; }}

/* Footer oculto. NO tocamos el header ni el toolbar: Streamlit usa esos
   contenedores para el botón de reabrir la sidebar cuando está colapsada,
   y cualquier regla sobre [data-testid="collapsedControl"] o el toolbar
   la rompe. */
footer {{ visibility: hidden; }}

/* Menú hamburguesa oculto sin tocar el resto del header */
#MainMenu {{ visibility: hidden; }}

/* Reducir el padding del contenedor principal para aprovechar más el ancho.
   Streamlit usa por defecto ~6rem de padding lateral, demasiado para una
   pantalla con sidebar lateral. Lo bajamos para que las tarjetas KPI no
   queden estrechas y los gráficos respiren mejor. */
.main .block-container,
[data-testid="stMain"] .block-container,
[data-testid="stAppViewContainer"] .block-container {{
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    padding-top: 2rem !important;
    max-width: 100% !important;
}}

/* Sidebar */
[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, var(--blue-dark) 0%, #00284a 100%);
    border-right: 1px solid rgba(255,255,255,0.08);
}}
[data-testid="stSidebar"] * {{ color: #e9eef5 !important; }}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {{ color: #fff !important; font-family: '{FONT_HEADING}', sans-serif !important; }}
[data-testid="stSidebar"] label {{
    font-family: '{FONT_HEADING}', sans-serif !important;
    font-size: 0.72rem !important;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    color: #b9c6d6 !important;
    font-weight: 600 !important;
}}
[data-testid="stSidebar"] [data-baseweb="select"] > div {{
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.14) !important;
    color: #fff !important;
}}
[data-testid="stSidebar"] button {{
    background: var(--orange-deep) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 2px !important;
    font-family: '{FONT_HEADING}', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
    font-size: 0.75rem !important;
    transition: background 0.2s ease;
}}
[data-testid="stSidebar"] button:hover {{
    background: var(--amber) !important;
}}
[data-testid="stSidebar"] [data-testid="stFileUploader"] section {{
    background: rgba(255,255,255,0.04) !important;
    border: 1px dashed rgba(255,255,255,0.25) !important;
    border-radius: 2px !important;
}}
[data-testid="stSidebar"] [data-testid="stFileUploader"] small {{
    color: #b9c6d6 !important;
}}

/* KPIs */
[data-testid="stMetric"] {{
    background: #fff;
    border: 1px solid var(--hairline);
    border-left: 3px solid var(--blue);
    padding: 0.9rem 1rem;
    border-radius: 2px;
    box-shadow: 0 1px 0 rgba(13,27,42,0.03);
    overflow: hidden;
}}
[data-testid="stMetric"] [data-testid="stMetricLabel"] {{
    font-family: '{FONT_HEADING}', sans-serif !important;
    font-size: 0.68rem !important;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    color: var(--ink-mute) !important;
    font-weight: 600;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}}
[data-testid="stMetric"] [data-testid="stMetricValue"] {{
    font-family: '{FONT_DISPLAY}', '{FONT_HEADING}', Helvetica, sans-serif !important;
    font-weight: 400 !important;
    font-size: 1.7rem !important;
    color: var(--ink) !important;
    letter-spacing: -0.015em;
    line-height: 1.05;
    white-space: nowrap;
    overflow: visible;
}}
[data-testid="stMetric"] [data-testid="stMetricValue"] > div {{
    overflow: visible !important;
}}
[data-testid="stMetric"] [data-testid="stMetricDelta"] {{
    font-family: '{FONT_MONO}', monospace !important;
    font-size: 0.78rem !important;
    color: var(--green-light) !important;
}}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {{
    gap: 0;
    border-bottom: 1px solid var(--hairline);
    background: transparent;
}}
.stTabs [data-baseweb="tab"] {{
    height: 48px;
    padding: 0 1.3rem;
    background: transparent !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    border-radius: 0 !important;
    color: var(--ink-mute) !important;
    font-family: '{FONT_HEADING}', sans-serif !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    transition: color 0.2s ease, border-color 0.2s ease;
}}
.stTabs [data-baseweb="tab"]:hover {{
    color: var(--blue) !important;
}}
.stTabs [aria-selected="true"] {{
    color: var(--blue-dark) !important;
    border-bottom: 2px solid var(--orange-deep) !important;
    background: transparent !important;
}}

/* Toggle gráfico/tabla */
div[data-testid="stRadio"] > label {{
    font-family: '{FONT_HEADING}', sans-serif !important;
    font-size: 0.7rem !important;
    text-transform: uppercase;
    letter-spacing: 0.16em;
    color: var(--ink-mute) !important;
    font-weight: 600 !important;
}}

/* Tablas nativas de Streamlit */
.stDataFrame, [data-testid="stDataFrame"] {{
    border: 1px solid var(--hairline);
    border-radius: 2px;
}}

/* Alertas */
.stAlert {{
    border-radius: 2px !important;
    border-left: 3px solid var(--blue) !important;
}}

/* Selects */
[data-baseweb="select"] > div {{
    border-radius: 2px !important;
    border-color: var(--hairline) !important;
}}

/* Botón de descarga */
.stDownloadButton button {{
    background: var(--blue-dark) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 2px !important;
    font-family: '{FONT_HEADING}', sans-serif !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    font-size: 0.78rem !important;
    padding: 0.55rem 1.1rem !important;
}}
.stDownloadButton button:hover {{
    background: var(--blue) !important;
}}

/* Encabezado editorial */
.masthead {{
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    padding: 1.2rem 0 0.4rem 0;
    border-bottom: 1px solid var(--ink);
    margin-bottom: 0.4rem;
}}
.masthead .eyebrow {{
    font-family: '{FONT_MONO}', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.22em;
    text-transform: uppercase;
    color: var(--orange-deep);
    margin-bottom: 0.4rem;
}}
.masthead h1 {{
    font-family: '{FONT_DISPLAY}', '{FONT_HEADING}', Helvetica, sans-serif !important;
    font-size: 4rem !important;
    line-height: 0.95 !important;
    margin: 0 !important;
    font-weight: 400 !important;
    color: var(--ink);
}}
.masthead h1 em {{
    font-style: normal;
    color: var(--orange-deep);
}}
.masthead .edition {{
    font-family: '{FONT_MONO}', monospace;
    font-size: 0.72rem;
    color: var(--ink-mute);
    text-align: right;
    letter-spacing: 0.08em;
    line-height: 1.55;
}}
.masthead .edition strong {{
    color: var(--ink);
    font-weight: 700;
    font-family: '{FONT_HEADING}', sans-serif;
    letter-spacing: 0.06em;
}}

.subhead {{
    font-family: '{FONT_BODY}', sans-serif;
    font-size: 0.92rem;
    color: var(--ink-mute);
    font-style: italic;
    border-bottom: 1px solid var(--hairline);
    padding-bottom: 1.2rem;
    margin-bottom: 1.6rem;
    letter-spacing: 0.01em;
    line-height: 1.5;
}}

/* Títulos de sección numerados */
.section-title {{
    font-family: '{FONT_HEADING}', sans-serif;
    font-size: 1.7rem;
    font-weight: 700;
    color: var(--ink);
    margin: 1.2rem 0 0.2rem 0;
    letter-spacing: -0.015em;
}}
.section-title .num {{
    font-family: '{FONT_MONO}', monospace;
    font-size: 0.75rem;
    color: var(--orange-deep);
    letter-spacing: 0.2em;
    vertical-align: middle;
    margin-right: 0.8rem;
    font-weight: 500;
}}
.section-title .seccion-info {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 22px;
    height: 22px;
    border-radius: 50%;
    background: var(--blue-dark);
    color: #fff;
    font-family: '{FONT_HEADING}', sans-serif;
    font-size: 0.8rem;
    font-weight: 700;
    margin-left: 0.7rem;
    cursor: help;
    vertical-align: middle;
    transition: background 0.18s ease;
}}
.section-title .seccion-info:hover {{
    background: var(--orange-deep);
}}
.section-kicker {{
    font-family: '{FONT_BODY}', sans-serif;
    font-size: 0.83rem;
    color: var(--ink-mute);
    font-style: italic;
    margin-bottom: 1rem;
    border-bottom: 1px dotted var(--hairline);
    padding-bottom: 0.8rem;
}}

/* Separador */
hr {{
    border: none !important;
    border-top: 1px solid var(--hairline) !important;
    margin: 2rem 0 !important;
}}

/* Caption */
[data-testid="stCaptionContainer"], .stCaption {{
    color: var(--ink-mute) !important;
    font-style: italic;
    font-size: 0.82rem !important;
    font-family: '{FONT_BODY}', sans-serif !important;
}}

/* Plotly containers */
.js-plotly-plot {{
    border: 1px solid var(--hairline);
    border-radius: 2px;
    background: #fff;
    padding: 0.6rem;
}}

/* Radio del sidebar */
[data-testid="stSidebar"] [role="radiogroup"] label {{
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.1);
    padding: 0.5rem 0.75rem;
    border-radius: 2px;
    margin-bottom: 0.3rem;
    text-transform: none;
    letter-spacing: 0;
    font-size: 0.85rem !important;
}}

/* Multiselect chips */
[data-baseweb="tag"] {{
    background: var(--blue-dark) !important;
    color: #fff !important;
    border-radius: 2px !important;
    font-family: '{FONT_HEADING}', sans-serif !important;
}}

/* =====================================================================
   Tabla institucional custom (.institutional-table)
   ===================================================================== */
.institutional-table {{
    width: 100%;
    border-collapse: collapse;
    font-family: '{FONT_BODY}', sans-serif;
    font-size: 0.85rem;
    background: #fff;
    border: 1px solid var(--hairline);
    border-radius: 2px;
    overflow: hidden;
    margin: 0.3rem 0 1.2rem 0;
}}
.institutional-table thead tr {{
    background: var(--blue-dark);
}}
.institutional-table thead th {{
    font-family: '{FONT_HEADING}', sans-serif;
    color: #fff;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    font-size: 0.7rem;
    text-align: left;
    padding: 0.8rem 0.9rem;
    border-bottom: 2px solid var(--orange-deep);
    white-space: nowrap;
}}
.institutional-table thead th.num {{
    text-align: right;
}}
.institutional-table tbody td {{
    padding: 0.65rem 0.9rem;
    border-bottom: 1px solid var(--hairline);
    color: var(--ink);
    vertical-align: middle;
}}
.institutional-table tbody td.num {{
    text-align: right;
    font-variant-numeric: tabular-nums;
    font-family: '{FONT_MONO}', monospace;
    font-size: 0.82rem;
    color: var(--ink);
}}
.institutional-table tbody tr:nth-child(even) {{
    background: #f6f6f5;
}}
.institutional-table tbody tr:hover {{
    background: #ededeb;
}}
.institutional-table tbody tr:last-child td {{
    border-bottom: none;
}}
.institutional-table tfoot td {{
    font-family: '{FONT_HEADING}', sans-serif;
    font-weight: 700;
    background: #ededeb;
    padding: 0.75rem 0.9rem;
    border-top: 2px solid var(--blue-dark);
    color: var(--blue-dark);
    text-transform: uppercase;
    letter-spacing: 0.06em;
    font-size: 0.78rem;
}}
.institutional-table tfoot td.num {{
    font-family: '{FONT_MONO}', monospace;
    font-size: 0.82rem;
    text-align: right;
    color: var(--blue-dark);
}}

/* Barra de progreso inline en celdas de porcentaje */
.pct-cell {{
    display: flex;
    align-items: center;
    gap: 0.6rem;
    justify-content: flex-end;
}}
.pct-cell .bar {{
    position: relative;
    width: 60px;
    height: 6px;
    background: var(--chip-bg);
    border-radius: 1px;
    overflow: hidden;
    flex-shrink: 0;
}}
.pct-cell .bar > span {{
    position: absolute;
    left: 0;
    top: 0;
    bottom: 0;
    background: var(--blue);
    border-radius: 1px;
}}
.pct-cell.low    .bar > span {{ background: var(--coral); }}
.pct-cell.mid    .bar > span {{ background: var(--amber); }}
.pct-cell.high   .bar > span {{ background: var(--green-light); }}
.pct-cell.top    .bar > span {{ background: var(--green-dark); }}

.pct-cell .value {{
    font-family: '{FONT_MONO}', monospace;
    font-size: 0.82rem;
    min-width: 55px;
    text-align: right;
    color: var(--ink);
}}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# =========================================================================
# Tema Plotly corporativo
# =========================================================================
CORPORATE_SEQUENCE = [
    COLORS["blue_dark"], COLORS["orange_deep"], COLORS["green_light"],
    COLORS["cyan"], COLORS["brown"], COLORS["blue"],
    COLORS["coral"], COLORS["amber"], COLORS["green_dark"], COLORS["orange"],
]

SCALE_BLUE = [
    [0.0, "#e8eef6"], [0.25, "#a9bedb"],
    [0.5, "#5f85b8"], [0.75, COLORS["blue"]], [1.0, COLORS["blue_dark"]],
]
SCALE_ORANGE = [
    [0.0, "#fbecd4"], [0.25, "#f3c77a"],
    [0.5, COLORS["orange"]], [0.75, COLORS["orange_deep"]], [1.0, COLORS["brown"]],
]
SCALE_GREEN = [
    [0.0, "#e1eee4"], [0.25, "#8ebfa0"],
    [0.5, COLORS["green_light"]], [0.75, COLORS["green_dark"]], [1.0, "#003d22"],
]

corporate_template = go.layout.Template()
corporate_template.layout = go.Layout(
    font=dict(family=f"{FONT_BODY}, sans-serif", color=COLORS["blue_dark"], size=12),
    title=dict(font=dict(family=f"{FONT_HEADING}, sans-serif", size=15, color="#0d1b2a")),
    paper_bgcolor="white",
    plot_bgcolor="white",
    colorway=CORPORATE_SEQUENCE,
    xaxis=dict(
        gridcolor="#ececea", linecolor="#c8c8c5", zerolinecolor="#ececea",
        ticks="outside", tickfont=dict(size=11, color="#4a5a6a"),
        title=dict(font=dict(size=11, color="#4a5a6a")),
    ),
    yaxis=dict(
        gridcolor="#ececea", linecolor="#c8c8c5", zerolinecolor="#ececea",
        ticks="outside", tickfont=dict(size=11, color="#4a5a6a"),
        title=dict(font=dict(size=11, color="#4a5a6a")),
    ),
    legend=dict(
        bgcolor="rgba(0,0,0,0)", bordercolor="#d9d4c7", borderwidth=1,
        font=dict(size=11, color="#0d1b2a"),
    ),
    margin=dict(l=60, r=30, t=60, b=60),
)
pio.templates["corporate"] = corporate_template
pio.templates.default = "corporate"


# =========================================================================
# URLs GitHub (datos que siempre vienen del repo)
# =========================================================================
GH = {
    "pi":  "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/Plan%20Indicativo%202024-2027.xlsx",
    "h24": "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/EJECUCION%20INVERSION%20A%20DICIEMBRE%2031%20DEL%202024%20ENERO%2010%202025.xlsx",
    "r24": "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/INFORME%20FINANCIERO%20REGALIAS%20A%2031%20DE%20DICIEMBRE%20DE%202024.xlsx",
    "h25": "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/EJECUCION%20INVERSION%20DE%20ENERO%20A%20DICIEMBRE%202025.xlsx",
    "r25": "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/PAGOS%20REGALIAS%20ENERO%20-%20DICIEMBRE%202025.xlsx",
    # Fuentes adicionales 2025 (vigencia cerrada)
    "ads_rp_25":     "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/RELACION%20DE%20PAGOS%20ENERO%20A%20DICIEMBRE%20ADS.xlsx",
    "ads_reg_25":    "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/PAGOS%20REGALIAS%202025%20ADS.xlsx",
    "gestiones_25":  "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/EjecucionFinancieraGestiones_20260210.xlsx",
    "fondo_mixto_25":"https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/CONTRATOS%20Y%20CONVENIOS%202025%20-%20FONDO%20MIXTO.xlsx",
    "inder_25":      "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/EjecucionIndersucre_Territorial_Regalias_202602010.xlsx",
    # Archivos actualizables
    "h26": "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/EJECUCION%20INVERSION%20DE%20HACIENDA%20PRUEBA%202026.xlsx",
    "r26": "https://raw.githubusercontent.com/Dona121/Plan-Indicativo/main/data/CG-cttos_04_marzo_20260304.xlsx",
}

# Vigencias cerradas (siempre se descargan del repo)
ARCHIVOS_CERRADOS = [
    "h24", "r24", "h25", "r25",
    "ads_rp_25", "ads_reg_25", "gestiones_25", "fondo_mixto_25", "inder_25",
]
# Archivos que el usuario puede actualizar manualmente
ARCHIVOS_ACTUALIZABLES = ["pi", "h26", "r26"]

# =========================================================================
# Utilidades
# =========================================================================
# Descarga desde GitHub
# - ttl=3600: la caché expira automáticamente cada hora, así que si subes un
#   archivo nuevo al repo basta con esperar (o usar el botón "Recargar datos").
# - El cache-buster `_t=...` se añade en la URL al pulsar el botón para
#   forzar bypass del CDN de GitHub raw.
# =========================================================================
@st.cache_data(show_spinner=False, ttl=3600)
def descargar_desde_github(url: str) -> bytes:
    # Si la URL trae el cache-buster, lo enviamos en el query string para
    # saltarnos cualquier capa de CDN o proxy.
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    return r.content


def formato_pesos(valor):
    try:
        v = float(valor)
        if abs(v) >= 1e9:
            return f"$ {v/1e9:,.2f} MM"
        if abs(v) >= 1e6:
            return f"$ {v/1e6:,.1f} M"
        return f"$ {v:,.0f}"
    except Exception:
        return valor


def formato_porcentaje(valor):
    try:
        return f"{valor*100:,.2f}%"
    except Exception:
        return valor


def formato_pesos_completo(valor):
    """Formato completo $ xxx,xxx,xxx para las tablas."""
    try:
        v = float(valor)
        return f"$ {v:,.0f}"
    except Exception:
        return valor


def formato_entero(valor):
    try:
        return f"{int(valor):,}"
    except Exception:
        return valor


def seccion(numero: str, titulo: str, kicker: str = "", tooltip: str = ""):
    """Encabezado de sección numerado.

    Si se pasa `tooltip`, aparece un ícono '?' al lado del título que muestra
    la explicación detallada al pasar el mouse encima.
    """
    icono = ""
    if tooltip:
        # El atributo title de HTML produce un tooltip nativo del navegador,
        # ligero y consistente con la estética minimalista del dashboard.
        tooltip_safe = tooltip.replace('"', "&quot;")
        icono = (
            f'<span class="seccion-info" title="{tooltip_safe}">?</span>'
        )
    st.markdown(
        f'<div class="section-title"><span class="num">{numero}</span>{titulo}{icono}</div>',
        unsafe_allow_html=True,
    )
    if kicker:
        st.markdown(f'<div class="section-kicker">{kicker}</div>', unsafe_allow_html=True)


# =========================================================================
# TOOLTIPS — explicaciones de cómo se calcula cada métrica
# =========================================================================
# Diccionario centralizado: cada clave corresponde a una métrica visible en
# la app. El usuario verá estos textos al pasar el mouse sobre el ícono "?".
TOOLTIPS = {
    # KPIs de cabecera
    "prog_vigencia": (
        "Total de recursos que el Plan Indicativo tiene presupuestados para la "
        "vigencia, sumando todas las fuentes de financiación: recursos propios "
        "(ICLD e ICDE), Sistema General de Participaciones (Educación, Salud y "
        "APSB), Regalías, cofinanciación de la Nación y de Municipios, crédito "
        "y otras fuentes."
    ),
    "ejec_vigencia": (
        "Recursos efectivamente pagados durante la vigencia. Consolida la "
        "información de Hacienda, Regalías y entidades adscritas. Para 2025 "
        "se incluyen además Aguas de Sucre, Gestiones, PDET, Fondo Mixto e "
        "Indersucre, que reportan su ejecución por separado."
    ),
    "avance_vigencia": (
        "Qué porcentaje del presupuesto programado se ha pagado efectivamente "
        "en la vigencia. Es la relación entre lo ejecutado y lo programado."
    ),
    "prog_cuatrienio": (
        "Recursos totales que el Plan Indicativo proyecta invertir en los "
        "cuatro años del Plan de Desarrollo (2024-2027), considerando todas "
        "las fuentes de financiación."
    ),
    "ejec_acumulada": (
        "Suma de los recursos pagados desde el inicio del Plan hasta la "
        "vigencia actual. Acumula la ejecución de 2024, 2025 y 2026. No "
        "incluye 2027 porque aún no ha iniciado."
    ),
    "avance_cuatrienio": (
        "Qué porcentaje del Plan de Desarrollo se ha ejecutado financieramente "
        "hasta el momento. Compara la ejecución acumulada contra la "
        "programación total del cuatrienio."
    ),

    # Avances físicos
    "avance_vig_ponderado": (
        "Mide qué tanto se han cumplido las metas del Plan en la vigencia. "
        "Es un promedio ponderado donde cada programa aporta según el número "
        "de metas que tiene programadas; los programas con más metas pesan "
        "proporcionalmente más en el resultado global."
    ),
    "avance_cuatrienio_ponderado": (
        "Cumplimiento global de las metas físicas del Plan considerando todo "
        "el cuatrienio, no solo la vigencia actual. Cada programa aporta "
        "según el peso que tiene dentro del Plan."
    ),
    "eficacia_operativa": (
        "Mide qué tan eficientes son las líneas y sectores en cumplir sus "
        "metas, ajustando por su tamaño relativo. Permite comparar de forma "
        "justa dependencias con muchas metas frente a otras con pocas: una "
        "línea con pocas metas pero alto cumplimiento puede tener mejor "
        "eficacia operativa que otra con muchas metas y bajo cumplimiento."
    ),
    "aporte_pdd": (
        "Cuánto contribuye una Línea Estratégica o un Sector al cumplimiento "
        "global del Plan de Desarrollo. Combina el nivel de avance de la "
        "agrupación con su peso dentro del total de metas."
    ),

    # Tablas financieras
    "ejec_clasif_recursos": (
        "Recursos pagados durante la vigencia, agrupados por el tipo de "
        "fuente que los financia (ICLD, ICDE, SGP, Regalías, Cofinanciación, "
        "Crédito u Otras Fuentes)."
    ),
    "porcentaje_ejecucion_financiera": (
        "Qué tanto se ha utilizado cada tipo de recurso en relación con lo "
        "que se tenía programado."
    ),

    # Distribución
    "distribucion_metas": (
        "Indica cómo está repartida la programación física del Plan entre "
        "los cuatro años. Compara cuánto se planea cumplir en cada vigencia "
        "frente a la meta total del cuatrienio."
    ),

    # Dependencia
    "metas_programadas": (
        "Cantidad de metas físicas que la dependencia tiene asignadas y para "
        "las cuales hay un valor a cumplir en la vigencia."
    ),
    "metas_cumplidas_100": (
        "Metas que alcanzaron o superaron el 100% del avance esperado en la "
        "vigencia (categoría 'Superior')."
    ),
    "porcentaje_ejec_dependencia": (
        "Avance promedio de las metas asignadas a la dependencia en la "
        "vigencia. Considera solo las metas con programación para ese año."
    ),
    "porcentaje_ejec_acumulada_dependencia": (
        "Avance promedio acumulado (cuatrienio) de todas las metas asignadas "
        "a la dependencia, considerando los años transcurridos del Plan."
    ),

    # Proyectos
    "total_proyectos_gestiones": (
        "Cantidad de proyectos y gestiones registrados en el Plan Indicativo "
        "para la vigencia. Incluye iniciativas tanto de ejecución directa "
        "como de gestión de recursos."
    ),
    "avance_proyecto": (
        "Porcentaje de cumplimiento físico del proyecto: qué tanto se ha "
        "ejecutado frente a la meta planeada (en unidades físicas, no en pesos)."
    ),
}


def programacion_financiera(vigencia: str):
    return (
        pl.col("programación recursos propios icld" + vigencia)
        + pl.col("programación recursos propios icde" + vigencia)
        + pl.col("programación sgp educación" + vigencia)
        + pl.col("programación sgp salud" + vigencia)
        + pl.col("programación sgp apsb" + vigencia)
        + pl.col("programación cofinanciación municipio" + vigencia)
        + pl.col("programación cofinanciación nación" + vigencia)
        + pl.col("programación crédito" + vigencia)
        + pl.col("programación regalías" + vigencia)
        + pl.col("programación otras fuentes" + vigencia)
    )


# =========================================================================
# Renderizador de tabla institucional
# =========================================================================
def pct_class(valor: float) -> str:
    if pd.isna(valor) or valor is None:
        return ""
    v = float(valor)
    if v < 0.25:  return "low"
    if v < 0.5:   return "mid"
    if v < 0.85:  return "high"
    return "top"


def formato_numero_decimal(valor, decimales: int = 2):
    try:
        v = float(valor)
        # Si es entero sin decimales significativos, muestra sin decimales
        if v == int(v):
            return f"{int(v):,}"
        return f"{v:,.{decimales}f}"
    except Exception:
        return valor


def render_table(df: pd.DataFrame, columnas: list, totales: dict = None):
    """
    Render de tabla institucional en HTML.
    `columnas` es una lista de dict: {'key', 'label', 'type'}
        type ∈ {'text', 'money', 'pct', 'int', 'num2', 'pctbar'}
    `totales` es un dict opcional {key: (value, type)} para la fila de total.
    """
    html = ['<table class="institutional-table">']

    # Encabezado
    html.append("<thead><tr>")
    for col in columnas:
        cls = "num" if col["type"] in ("money", "pct", "int", "num2", "pctbar") else ""
        html.append(f'<th class="{cls}">{col["label"]}</th>')
    html.append("</tr></thead>")

    # Cuerpo
    html.append("<tbody>")
    for _, row in df.iterrows():
        html.append("<tr>")
        for col in columnas:
            v = row.get(col["key"])
            t = col["type"]
            if t == "money":
                cell = formato_pesos_completo(v) if pd.notna(v) else "—"
                html.append(f'<td class="num">{cell}</td>')
            elif t == "pct":
                cell = formato_porcentaje(v) if pd.notna(v) else "—"
                html.append(f'<td class="num">{cell}</td>')
            elif t == "pctbar":
                if pd.notna(v):
                    pct = max(0.0, min(1.0, float(v)))
                    klass = pct_class(v)
                    bar = int(pct * 100)
                    cell = (
                        f'<div class="pct-cell {klass}">'
                        f'<div class="bar"><span style="width:{bar}%"></span></div>'
                        f'<div class="value">{v*100:.1f}%</div>'
                        f'</div>'
                    )
                else:
                    cell = '<div class="pct-cell"><div class="value">—</div></div>'
                html.append(f'<td class="num">{cell}</td>')
            elif t == "int":
                cell = formato_entero(v) if pd.notna(v) else "—"
                html.append(f'<td class="num">{cell}</td>')
            elif t == "num2":
                cell = formato_numero_decimal(v) if pd.notna(v) else "—"
                html.append(f'<td class="num">{cell}</td>')
            else:
                cell = "" if pd.isna(v) or v is None else str(v)
                html.append(f"<td>{cell}</td>")
        html.append("</tr>")
    html.append("</tbody>")

    # Totales
    if totales:
        html.append("<tfoot><tr>")
        for i, col in enumerate(columnas):
            if i == 0 and col["key"] not in totales:
                html.append("<td>Total</td>")
                continue
            v = totales.get(col["key"])
            if v is None:
                cls = "num" if col["type"] in ("money", "pct", "int", "num2", "pctbar") else ""
                html.append(f'<td class="{cls}"></td>')
                continue
            t = col["type"]
            if t == "money":
                html.append(f'<td class="num">{formato_pesos_completo(v)}</td>')
            elif t == "pct" or t == "pctbar":
                html.append(f'<td class="num">{formato_porcentaje(v)}</td>')
            elif t == "int":
                html.append(f'<td class="num">{formato_entero(v)}</td>')
            elif t == "num2":
                html.append(f'<td class="num">{formato_numero_decimal(v)}</td>')
            else:
                html.append(f"<td>{v}</td>")
        html.append("</tr></tfoot>")

    html.append("</table>")
    st.markdown("".join(html), unsafe_allow_html=True)


def render_vista(
    tipo_vista: str,
    fig_factory,
    df_tabla: pd.DataFrame,
    columnas: list,
    totales: dict = None,
):
    """Renderiza gráfico o tabla según la selección del usuario."""
    if tipo_vista == "Tabla":
        render_table(df_tabla, columnas, totales)
    else:
        st.plotly_chart(fig_factory(), use_container_width=True)


def selector_vista(key: str) -> str:
    return st.radio(
        "Vista",
        options=["Gráfico", "Tabla"],
        horizontal=True,
        key=key,
        label_visibility="collapsed",
    )


# =========================================================================
# Procesamiento
# =========================================================================
@st.cache_data(show_spinner="Procesando datos del Plan Indicativo...")
def procesar_datos(
    pi_bytes, h24_bytes, r24_bytes, h25_bytes, r25_bytes,
    ads_rp_25_bytes, ads_reg_25_bytes, gestiones_25_bytes,
    fondo_mixto_25_bytes, inder_25_bytes,
    h26_bytes, r26_bytes,
):
    plan_indicativo = pl.read_excel(io.BytesIO(pi_bytes), table_name="tblPlanIndicativo_2")
    orden_lineas_pdd = pl.read_excel(io.BytesIO(pi_bytes), table_name="orden_lineas")
    orden_sectores_pdd = pl.read_excel(io.BytesIO(pi_bytes), table_name="orden_sectores")
    orden_programas_pdd = pl.read_excel(io.BytesIO(pi_bytes), table_name="orden_programas")
    homologacion_secretarias = pl.read_excel(io.BytesIO(pi_bytes), table_name="HomologacionSecretarias")

    columnas_prog_ejec_fisica = plan_indicativo.select(
        "Codigo Meta", "Línea Estratégica", "Sector PDD",
        "Numero Programa PDD", "Programa PDD",
        "Indicador de producto principal", "código de indicador principal",
        "Meta de cuatrenio",
        "Tipo de Acumulación", "Responsable", "Meta Física Esperada 2024",
        "Meta Física Esperada 2025", "Meta Física Esperada 2026", "Meta Física Esperada 2027",
        "PROYECTOS 2024", "PROYECTOS 2025", "PROYECTOS/GESTIONES PROGRAMADAS 2026", "PROYECTOS 2026",
        "PROYECTOS 2027", "EJECUCIÓN 2024", "PORCENTAJE DE EJECUCIÓN 2024", "CATEGORÍA DE EJECUCIÓN FÍSICA 2024",
        "EJECUCIÓN 2025", "PORCENTAJE DE EJECUCIÓN 2025", "CATEGORÍA DE EJECUCIÓN FÍSICA 2025",
        "EJECUCIÓN 2026", "PORCENTAJE DE EJECUCIÓN 2026", "CATEGORÍA DE EJECUCIÓN FÍSICA 2026",
        "EJECUCIÓN ACUMULADA", "PORCENTAJE DE EJECUCIÓN ACUMULADA", "CATEGORÍA DE EJECUCIÓN ACUMULADA",
    )

    # --- Ejecución 2024 ---
    ejecucion_regalias_2024 = (
        pl.read_excel(io.BytesIO(r24_bytes), table_name="EjecucionRegalias",
                      columns=["CODIGO META", "COMPROMISOS", "CLASIFICACIÓN RECURSOS"])
        .with_columns(pl.col("CODIGO META").fill_null(pl.lit("")))
        .filter(pl.col("CODIGO META") != "", pl.col("CODIGO META").str.starts_with("MT"))
        .rename({"COMPROMISOS": "RP"})
    )
    ejecucion_hacienda_2024 = (
        pl.read_excel(io.BytesIO(h24_bytes), table_name="EjecucionHaciendaDiciembre",
                      columns=["RP", "CODIGO META", "CLASIFICACIÓN RECURSOS"])
        .with_columns(pl.col("CODIGO META", "CLASIFICACIÓN RECURSOS").fill_null(pl.lit("")))
        .filter(pl.col("CODIGO META") != "", pl.col("CLASIFICACIÓN RECURSOS") != "")
    )

    # --- Ejecución 2025 (fuentes base) ---
    ejecucion_regalias_2025 = (
        pl.read_excel(io.BytesIO(r25_bytes), table_name="Pagos_Regalias_2025")
        .select("PAGOS REGALIAS", "CODIGO META", "CLASIFICACIÓN RECURSOS")
        .rename({"PAGOS REGALIAS": "RP"})
        .with_columns(pl.col("CODIGO META").fill_null(pl.lit("")))
        .filter(pl.col("CODIGO META") != "")
    )
    ejecucion_hacienda_2025 = (
        pl.read_excel(io.BytesIO(h25_bytes), table_name="EjecucionHaciendaDiciembre2025")
        .with_columns(
            pl.col("PROYECTO ARCHIVADO", "CODIGO META", "CLASIFICACIÓN RECURSOS", "SE VA A CARGAR EN PI").fill_null(pl.lit("")),
            pl.when(pl.col("DISTRIBUIR DE FORMA EQUITATIVA") == "SI").then(pl.col("RP") / 2).otherwise(pl.col("RP")),
        )
        .filter(pl.col("PROYECTO ARCHIVADO") == "", pl.col("CODIGO META") != "",
                pl.col("CLASIFICACIÓN RECURSOS") != "", pl.col("SE VA A CARGAR EN PI") == "")
        .select("CODIGO META", "CLASIFICACIÓN RECURSOS", "RP")
    )

    # --- Ejecución 2025 (fuentes adicionales) ---
    ejecucion_2025_ads_recursos_propios = (
        pl.read_excel(io.BytesIO(ads_rp_25_bytes), table_name="PagosAguasDeSucre")
        .select("VALOR DEL PAGO", "CLASIFICACIÓN RECURSOS", "CODIGO META")
        .with_columns(pl.col("CODIGO META").fill_null(pl.lit("")))
        .filter(pl.col("CODIGO META") != "")
        .rename({"VALOR DEL PAGO": "RP"})
    )

    ejecucion_2025_ads_regalias = (
        pl.read_excel(io.BytesIO(ads_reg_25_bytes), table_name="RegaliasAguasDeSucre")
        .select("CODIGO DE META", "CLASIFICACIÓN RECURSOS", "PAGOS")
        .rename({"CODIGO DE META": "CODIGO META", "PAGOS": "RP"})
        .with_columns(pl.col("CODIGO META").fill_null(pl.lit("")))
        .filter(pl.col("CODIGO META") != "")
    )

    ejecucion_2025_gestiones = (
        pl.read_excel(io.BytesIO(gestiones_25_bytes), table_name="EjecucionGestiones")
        .rename({"EJECUCION FINANCIERA": "RP"})
    )

    ejecucion_pdet_2025 = (
        pl.read_excel(io.BytesIO(h25_bytes), table_name="EjecucionPDET")
        .select("EJECUCION FINANCIERA", "CODIGO META", "CLASIFICACIÓN RECURSOS")
        .rename({"EJECUCION FINANCIERA": "RP"})
    )

    ejecucion_2025_fondo_mixto = (
        pl.read_excel(io.BytesIO(fondo_mixto_25_bytes), table_name="EjecucionFinancieraFondoMixto")
        .select("CLASIFICACIÓN RECURSOS", "EJECUCION FINANCIERA", "CODIGO META")
        .rename({"EJECUCION FINANCIERA": "RP"})
    )

    ejecucion_2025_indersucre_recursos_propios = (
        pl.read_excel(io.BytesIO(inder_25_bytes), table_name="EjecucionFinancieraINDERTerritorio")
        .rename({"EJECUCION FINANCIERA": "RP"})
    )

    ejecucion_2025_indersucre_regalias = (
        pl.read_excel(io.BytesIO(inder_25_bytes), table_name="EjecucionFinancieraINDERRegalias")
        .rename({"EJECUCION FINANCIERA": "RP"})
    )

    # --- Ejecución 2026 ---
    ejecucion_regalias_2026 = (
        pl.read_excel(io.BytesIO(r26_bytes), table_name="Pagos_Regalias_2026")
        .select(pl.all().name.map(lambda x: x.strip().upper().replace("_X0009_", "")))
        .filter(
            (pl.col("ULTIMA FECHA PAGO") >= pl.date(2026, 1, 1))
            & (pl.col("ULTIMA FECHA PAGO") <= pl.date(2026, 12, 31))
        )
        .select("PAGO EJECUTADO VALOR", "CODIGO META", "CLASIFICACIÓN RECURSOS")
        .rename({"PAGO EJECUTADO VALOR": "RP"})
        .with_columns(pl.col("CODIGO META").fill_null(pl.lit("")))
        .filter(pl.col("CODIGO META") != "")
    )
    ejecucion_hacienda_2026 = (
        pl.read_excel(io.BytesIO(h26_bytes), table_name="EjecucionHacienda2026")
        .with_columns(
            pl.col("PROYECTO ARCHIVADO", "CODIGO META", "CLASIFICACIÓN RECURSOS", "SE VA A CARGAR EN PI").fill_null(pl.lit("")),
            pl.when(pl.col("DISTRIBUIR DE FORMA EQUITATIVA") == "SI").then(pl.col("RP") / 2).otherwise(pl.col("RP")),
        )
        .filter(pl.col("PROYECTO ARCHIVADO") == "", pl.col("CODIGO META") != "",
                pl.col("CLASIFICACIÓN RECURSOS") != "", pl.col("SE VA A CARGAR EN PI") == "")
        .select("CODIGO META", "CLASIFICACIÓN RECURSOS", "RP")
    )

    # --- Agregados de ejecución por meta ---
    ejec_2024 = pl.concat([ejecucion_regalias_2024, ejecucion_hacienda_2024], how="diagonal") \
        .group_by("CODIGO META").agg(pl.col("RP").sum().alias("Ejecución Financiera 2024"))

    # 2025: las fuentes adicionales pueden traer códigos múltiples separados por " | ".
    # El notebook hace str.split(" | ").explode() SIN dividir el RP, por lo que cada
    # meta de un código múltiple recibe el RP completo de la fila original.
    # Esto puede inflar el total cuando se suma sobre todas las metas — pero es lo
    # que hace el notebook y la app debe reflejarlo fielmente.
    ejec_2025_full = (
        pl.concat([
            ejecucion_regalias_2025, ejecucion_hacienda_2025,
            ejecucion_2025_ads_recursos_propios, ejecucion_2025_ads_regalias,
            ejecucion_2025_gestiones, ejecucion_pdet_2025,
            ejecucion_2025_fondo_mixto,
            ejecucion_2025_indersucre_recursos_propios,
            ejecucion_2025_indersucre_regalias,
        ], how="diagonal")
        .with_columns(pl.col("CODIGO META").str.split(" | "))
        .explode("CODIGO META")
    )
    ejec_2025 = ejec_2025_full.group_by("CODIGO META").agg(
        pl.col("RP").sum().alias("Ejecución Financiera 2025")
    )

    ejec_2026 = pl.concat([ejecucion_regalias_2026, ejecucion_hacienda_2026], how="diagonal") \
        .group_by("CODIGO META").agg(pl.col("RP").sum().alias("Ejecución Financiera 2026"))

    columnas_programacion_financiera = (
        plan_indicativo.select("Codigo Meta", cs.starts_with("Programación").cast(pl.Float64))
        .select(pl.all().name.map(lambda x: x.strip().lower()))
        .select(
            "codigo meta",
            programacion_financiera("24").alias("Programación Financiera 2024"),
            programacion_financiera("25").alias("Programación Financiera 2025"),
            programacion_financiera("26").alias("Programación Financiera 2026"),
            programacion_financiera("27").alias("Programación Financiera 2027"),
        )
        .join(ejec_2024, left_on="codigo meta", right_on="CODIGO META", how="left")
        .join(ejec_2025, left_on="codigo meta", right_on="CODIGO META", how="left")
        .join(ejec_2026, left_on="codigo meta", right_on="CODIGO META", how="left")
        .with_columns(pl.col("Ejecución Financiera 2024", "Ejecución Financiera 2025", "Ejecución Financiera 2026").fill_null(pl.lit(0)))
    )

    orden_fuentes = pl.DataFrame({
        "Clasificación Recursos": ["COFINANCIACIÓN MUNICIPIO", "ICDE", "OTRAS FUENTES", "SGP APSB", "SGP SALUD",
                                   "SGP EDUCACION", "REGALÍAS", "COFINANCIACIÓN NACIÓN", "ICLD", "CREDITO"],
        "Orden": [1, 5, 3, 7, 9, 8, 10, 2, 6, 4],
        "Tipo Fuente": ["Otras Fuentes", "Recursos Propios", "Otras Fuentes",
                        "Sistema General de Participaciones (SGP)", "Sistema General de Participaciones (SGP)",
                        "Sistema General de Participaciones (SGP)", "Sistema General de Regalías",
                        "Otras Fuentes", "Recursos Propios", "Recursos del Crédito"],
    })

    prog_financ_tipo = (
        plan_indicativo.select(cs.starts_with("Programación"))
        .select(pl.all().name.map(lambda x: x.strip().lower()))
        .select(cs.exclude("programación total24", "programación total25", "programación total26", "programación total27"))
        .unpivot(on=cs.numeric(), variable_name="Clasificación Recursos", value_name="Programación financiera")
        .group_by("Clasificación Recursos").agg(pl.col("Programación financiera").sum())
        .with_columns(
            pl.col("Clasificación Recursos").str.slice(-2).alias("Vigencia"),
            pl.col("Clasificación Recursos").str.replace_all("programación ", "").str.replace_all(r"(24|25|26|27)", ""),
        )
        .with_columns((pl.lit("Programación Financiera 20") + pl.col("Vigencia")).alias("Vigencia"))
        .pivot(index="Clasificación Recursos", on="Vigencia", aggregate_function="sum")
        .with_columns(
            pl.col("Clasificación Recursos").str.replace_many(
                ["recursos propios icde", "cofinanciación nación", "sgp educación", "cofinanciación municipio",
                 "sgp salud", "sgp apsb", "otras fuentes", "regalías", "recursos propios icld", "crédito"],
                ["ICDE", "COFINANCIACIÓN NACIÓN", "SGP EDUCACION", "COFINANCIACIÓN MUNICIPIO",
                 "SGP SALUD", "SGP APSB", "OTRAS FUENTES", "REGALÍAS", "ICLD", "CREDITO"],
            )
        )
    )

    ejecuciones_financieras = {
        "2024": [ejecucion_regalias_2024, ejecucion_hacienda_2024],
        "2025": [
            ejecucion_regalias_2025, ejecucion_hacienda_2025,
            ejecucion_2025_ads_recursos_propios, ejecucion_2025_ads_regalias,
            ejecucion_2025_gestiones, ejecucion_pdet_2025,
            ejecucion_2025_fondo_mixto,
            ejecucion_2025_indersucre_recursos_propios,
            ejecucion_2025_indersucre_regalias,
        ],
        "2026": [ejecucion_regalias_2026, ejecucion_hacienda_2026],
    }

    prog_fisica_financiera = (
        columnas_prog_ejec_fisica.join(columnas_programacion_financiera, left_on="Codigo Meta", right_on="codigo meta", how="left")
        .with_columns(
            pl.col("Meta Física Esperada 2024", "Meta Física Esperada 2025",
                   "Meta Física Esperada 2026", "Meta Física Esperada 2027").fill_null(pl.lit(0))
        )
    )

    return {
        "plan_indicativo": plan_indicativo,
        "orden_lineas_pdd": orden_lineas_pdd,
        "orden_sectores_pdd": orden_sectores_pdd,
        "orden_programas_pdd": orden_programas_pdd,
        "homologacion_secretarias": homologacion_secretarias,
        "orden_fuentes": orden_fuentes,
        "prog_financ_tipo": prog_financ_tipo,
        "ejecuciones_financieras": ejecuciones_financieras,
        "prog_fisica_financiera": prog_fisica_financiera,
    }


# =========================================================================
# Constructores de reportes por vigencia
# =========================================================================
def construir_ejecucion_financ_tipo(datos, vigencia):
    """Ejecución por clasificación de recursos para la vigencia filtrada.

    En 2025 se aplica str.split(" | ").explode() ANTES de agrupar, igual que
    en el bloque acumulado. Esto garantiza que los KPIs y la tabla de
    'Por Clasificación de Recursos' reflejen el RP tras explode (la cifra
    real validada con el notebook).
    """
    ejecuciones = datos["ejecuciones_financieras"][vigencia]
    orden_fuentes = datos["orden_fuentes"]
    prog_financ_tipo = datos["prog_financ_tipo"]

    concat_ejec = pl.concat(ejecuciones, how="diagonal")
    if vigencia == "2025":
        concat_ejec = concat_ejec.with_columns(pl.col("CODIGO META").str.split(" | ")).explode("CODIGO META")

    return (
        orden_fuentes.join(
            concat_ejec,
            left_on="Clasificación Recursos", right_on="CLASIFICACIÓN RECURSOS", how="left",
        )
        .group_by("Clasificación Recursos").agg(pl.col("RP").sum().alias(f"Ejecución Financiera {vigencia}"))
        .join(orden_fuentes, on="Clasificación Recursos", how="inner")
        .join(prog_financ_tipo, on="Clasificación Recursos", how="inner")
        .select("Orden", "Tipo Fuente", "Clasificación Recursos",
                f"Programación Financiera {vigencia}", f"Ejecución Financiera {vigencia}")
        .with_columns(
            (pl.when(pl.col(f"Programación Financiera {vigencia}") == 0)
             .then(pl.lit(0))
             .otherwise(pl.col(f"Ejecución Financiera {vigencia}") / pl.col(f"Programación Financiera {vigencia}"))
             ).alias("Porcentaje de Ejecución Financiera")
        )
        .sort(by="Orden")
    )


def construir_ejecucion_acumulada_tipo(datos):
    """Acumulado por CLASIFICACIÓN RECURSOS para todo el cuatrienio.

    Replica exactamente la lógica del notebook (bloque ejecucion_2024_agrp,
    ejecucion_2025_agrp, ejecucion_2026_agrp). En 2025 se hace explode por
    " | " antes de agrupar — el notebook lo hace así y la app lo replica.
    """
    orden_fuentes = datos["orden_fuentes"]
    prog_financ_tipo = datos["prog_financ_tipo"]

    agrp = {}
    for v in ["2024", "2025", "2026"]:
        concat = pl.concat(datos["ejecuciones_financieras"][v], how="diagonal")
        if v == "2025":
            concat = concat.with_columns(pl.col("CODIGO META").str.split(" | ")).explode("CODIGO META")
        agrp[v] = (
            concat.group_by("CLASIFICACIÓN RECURSOS")
            .agg(pl.col("RP").sum().alias(f"Ejecución Financiera {v}"))
        )

    return (
        orden_fuentes
        .join(agrp["2024"], left_on="Clasificación Recursos", right_on="CLASIFICACIÓN RECURSOS", how="left")
        .join(agrp["2025"], left_on="Clasificación Recursos", right_on="CLASIFICACIÓN RECURSOS", how="left")
        .join(agrp["2026"], left_on="Clasificación Recursos", right_on="CLASIFICACIÓN RECURSOS", how="left")
        .join(prog_financ_tipo, on="Clasificación Recursos")
        .with_columns(
            pl.col("Ejecución Financiera 2024", "Ejecución Financiera 2025", "Ejecución Financiera 2026").fill_null(pl.lit(0))
        )
        .with_columns(
            (pl.col("Ejecución Financiera 2024") + pl.col("Ejecución Financiera 2025")
             + pl.col("Ejecución Financiera 2026")).alias("Ejecución Financiera Acumulada"),
            (pl.col("Programación Financiera 2024") + pl.col("Programación Financiera 2025")
             + pl.col("Programación Financiera 2026") + pl.col("Programación Financiera 2027")
             ).alias("Programación Cuatrienio"),
        )
        .select("Orden", "Tipo Fuente", "Clasificación Recursos",
                "Programación Financiera 2024", "Programación Financiera 2025", "Programación Financiera 2026",
                "Ejecución Financiera 2024", "Ejecución Financiera 2025", "Ejecución Financiera 2026",
                "Programación Cuatrienio", "Ejecución Financiera Acumulada")
        .sort(by="Orden")
    )


def construir_prog_financ_categorias(datos, vigencia):
    prog_ff = datos["prog_fisica_financiera"]
    orden_lineas = datos["orden_lineas_pdd"]
    orden_sectores = datos["orden_sectores_pdd"]
    orden_programas = datos["orden_programas_pdd"]

    def agregar(grupo, orden_df, col_orden):
        return (
            prog_ff.group_by(grupo).agg(
                pl.col(f"Programación Financiera {vigencia}").sum(),
                pl.col(f"Ejecución Financiera {vigencia}").sum(),
            )
            .join(orden_df, on=grupo, how="inner")
            .with_columns(
                (pl.when(pl.col(f"Programación Financiera {vigencia}") == 0)
                 .then(pl.lit(0))
                 .otherwise(pl.col(f"Ejecución Financiera {vigencia}") / pl.col(f"Programación Financiera {vigencia}"))
                 ).alias("Porcentaje de Ejecución Financiera")
            )
            .sort(col_orden)
            .select(col_orden, grupo, f"Programación Financiera {vigencia}",
                    f"Ejecución Financiera {vigencia}", "Porcentaje de Ejecución Financiera")
        )

    return {
        "lineas": agregar("Línea Estratégica", orden_lineas, "Orden Linea"),
        "sectores": agregar("Sector PDD", orden_sectores, "Orden Sector"),
        "programas": agregar("Programa PDD", orden_programas, "Orden Programa PDD"),
    }


def construir_ejec_por_dependencia(datos, vigencia):
    """Replica exactamente 'ejecucion_por_dependencia' del notebook.

    Usa join 'left' con la homologación de secretarías, igual que el notebook.
    El otro DataFrame del notebook ('avance_por_dependencia') usa inner pero
    no es el que alimenta esta tabla principal.
    """
    prog_ff = datos["prog_fisica_financiera"]
    homologacion = datos["homologacion_secretarias"]

    ejec_acumulada = (
        prog_ff.select(pl.col("Responsable").str.strip_chars(), "PORCENTAJE DE EJECUCIÓN ACUMULADA")
        .group_by("Responsable")
        .agg(pl.col("PORCENTAJE DE EJECUCIÓN ACUMULADA").fill_null(pl.lit(0)).mean().alias("Porcentaje de Ejecución Acumulada"))
    )

    return (
        prog_ff.select(
            pl.col("Responsable").str.strip_chars(), f"Meta Física Esperada {vigencia}",
            f"PORCENTAJE DE EJECUCIÓN {vigencia}", f"CATEGORÍA DE EJECUCIÓN FÍSICA {vigencia}"
        )
        .filter(pl.col(f"Meta Física Esperada {vigencia}").fill_null(pl.lit(0)) != 0)
        .with_columns(
            (pl.when(pl.col(f"Meta Física Esperada {vigencia}") != 0).then(pl.lit(1)).otherwise(pl.lit(0))
             ).alias(f"Metas Programadas {vigencia}"),
            (pl.when(pl.col(f"CATEGORÍA DE EJECUCIÓN FÍSICA {vigencia}") == "Superior").then(pl.lit(1)).otherwise(pl.lit(0))
             ).alias(f"Metas Cumplidas al 100% {vigencia}"),
        )
        .group_by("Responsable").agg(
            pl.col(f"PORCENTAJE DE EJECUCIÓN {vigencia}").fill_null(pl.lit(0)).mean().alias(f"Porcentaje de Ejecución {vigencia}"),
            pl.col(f"Metas Programadas {vigencia}").sum(),
            pl.col(f"Metas Cumplidas al 100% {vigencia}").sum(),
        )
        .join(homologacion, left_on="Responsable", right_on="Responsable en PI", how="left")
        .join(ejec_acumulada, on="Responsable", how="left")
        .select("Varias Secretarías", "Dependencia Responsable",
                f"Metas Programadas {vigencia}", f"Metas Cumplidas al 100% {vigencia}",
                f"Porcentaje de Ejecución {vigencia}", "Porcentaje de Ejecución Acumulada")
    )


def construir_avances_fisicos(datos, vigencia):
    prog_ff = datos["prog_fisica_financiera"]

    numero_total_metas = prog_ff.get_column("Codigo Meta").count()
    numero_metas_prog_vigencia = (
        prog_ff.filter(pl.col(f"Meta Física Esperada {vigencia}") != 0).get_column("Codigo Meta").count()
    )

    promedio_programas = (
        prog_ff.filter(pl.col(f"Meta Física Esperada {vigencia}") != 0)
        .group_by("Programa PDD").agg(pl.col(f"PORCENTAJE DE EJECUCIÓN {vigencia}").mean())
        .rename({f"PORCENTAJE DE EJECUCIÓN {vigencia}": "Promedio de avance de ejecución de la vigencia"})
    )

    num_metas_lineas_cp = (
        prog_ff.filter(pl.col(f"Meta Física Esperada {vigencia}") != 0)
        .group_by("Línea Estratégica").agg(pl.col("Codigo Meta").len())
        .rename({"Codigo Meta": "Total Indicadores de Producto con Programacion"})
    )
    num_metas_lineas = (
        prog_ff.group_by("Línea Estratégica").agg(pl.col("Codigo Meta").len())
        .rename({"Codigo Meta": "Total Indicadores de Producto"})
    )
    num_metas_sectores_cp = (
        prog_ff.filter(pl.col(f"Meta Física Esperada {vigencia}") != 0)
        .group_by("Sector PDD").agg(pl.col("Codigo Meta").len())
        .rename({"Codigo Meta": "Total Indicadores de Producto con Programacion"})
    )
    num_metas_sectores = (
        prog_ff.group_by("Sector PDD").agg(pl.col("Codigo Meta").len())
        .rename({"Codigo Meta": "Total Indicadores de Producto"})
    )
    num_metas_programas = (
        prog_ff.group_by("Programa PDD").agg(pl.col("Codigo Meta").len())
        .rename({"Codigo Meta": "Total Indicadores de Producto"})
    )

    ponderado_vigencia = (
        prog_ff
        .with_columns(
            (pl.when(pl.col(f"Meta Física Esperada {vigencia}") != 0).then(pl.lit(1)).otherwise(pl.lit(0))
             ).alias(f"Metas Programadas {vigencia}")
        )
        .group_by("Línea Estratégica", "Sector PDD", "Programa PDD")
        .agg(pl.col(f"Metas Programadas {vigencia}").sum())
        .with_columns(
            (pl.col(f"Metas Programadas {vigencia}") / max(numero_metas_prog_vigencia, 1)
             ).alias("Sobre Numero Total de Metas Programadas")
        )
        .join(promedio_programas, on="Programa PDD", how="left")
        .with_columns(pl.col("Promedio de avance de ejecución de la vigencia").fill_null(pl.lit(0)))
        .rename({f"Metas Programadas {vigencia}": "Total Indicadores de Producto Programados"})
    )

    ponderado_cuatrienio = (
        prog_ff.group_by("Línea Estratégica", "Sector PDD", "Programa PDD")
        .agg(pl.col("PORCENTAJE DE EJECUCIÓN ACUMULADA").fill_null(pl.lit(0)).mean())
        .join(num_metas_programas, on="Programa PDD")
        .with_columns((pl.col("Total Indicadores de Producto") / max(numero_total_metas, 1)).alias("Sobre Numero Total de Metas"))
        .rename({"PORCENTAJE DE EJECUCIÓN ACUMULADA": "Promedio de avance de ejecución acumulada"})
    )

    avance_vig_ponderado = ponderado_vigencia.select(
        pl.col("Promedio de avance de ejecución de la vigencia") * pl.col("Sobre Numero Total de Metas Programadas")
    ).sum().item()

    avance_cuatrienio_total = ponderado_cuatrienio.select(
        pl.col("Promedio de avance de ejecución acumulada") * pl.col("Sobre Numero Total de Metas")
    ).sum().item()

    def avance_por_dim(ponderado, grupo, num_metas_df, total, col_avance, col_metas):
        return (
            ponderado.group_by(grupo)
            .agg((pl.col(col_avance) * pl.col(col_metas)).sum())
            .rename({col_avance: "% Aporte Cumplimiento PDD"})
            .join(num_metas_df, on=grupo)
            .with_columns((pl.col(num_metas_df.columns[1]) / max(total, 1)).alias("Sobre Numero Total de Indicadores"))
            .with_columns(
                (pl.when(pl.col("Sobre Numero Total de Indicadores") == 0)
                 .then(pl.lit(0))
                 .otherwise(pl.col("% Aporte Cumplimiento PDD") / pl.col("Sobre Numero Total de Indicadores"))
                 ).alias("% Eficacia Operativa")
            )
        )

    return {
        "avance_vig_ponderado": avance_vig_ponderado,
        "avance_cuatrienio_total": avance_cuatrienio_total,
        "avance_vig_lineas": avance_por_dim(
            ponderado_vigencia, "Línea Estratégica", num_metas_lineas_cp, numero_metas_prog_vigencia,
            "Promedio de avance de ejecución de la vigencia", "Sobre Numero Total de Metas Programadas"),
        "avance_cuatri_lineas": avance_por_dim(
            ponderado_cuatrienio, "Línea Estratégica", num_metas_lineas, numero_total_metas,
            "Promedio de avance de ejecución acumulada", "Sobre Numero Total de Metas"),
        "avance_vig_sectores": avance_por_dim(
            ponderado_vigencia, "Sector PDD", num_metas_sectores_cp, numero_metas_prog_vigencia,
            "Promedio de avance de ejecución de la vigencia", "Sobre Numero Total de Metas Programadas"),
        "avance_cuatri_sectores": avance_por_dim(
            ponderado_cuatrienio, "Sector PDD", num_metas_sectores, numero_total_metas,
            "Promedio de avance de ejecución acumulada", "Sobre Numero Total de Metas"),
        "numero_total_metas": numero_total_metas,
        "numero_metas_prog_vigencia": numero_metas_prog_vigencia,
    }


# =========================================================================
# Constructor de proyectos por vigencia
# =========================================================================
def _extraer_regex(expr: pl.Expr, patron: str) -> pl.Expr:
    return expr.str.extract(patron, group_index=1).str.strip_chars()


def _normalizar_numero(expr: pl.Expr) -> pl.Expr:
    x = expr.str.strip_chars().str.replace_all(r"\s+", "")
    return (
        pl.when(x.is_null() | (x == ""))
        .then(pl.lit(None))
        .when(x.str.contains(r"^\d{1,3}(?:\.\d{3})+,\d+$"))
        .then(x.str.replace_all(r"\.", "").str.replace_all(",", "."))
        .when(x.str.contains(r"^\d{1,3}(?:,\d{3})+\.\d+$"))
        .then(x.str.replace_all(",", ""))
        .when(x.str.contains(r"^\d{1,3}(?:\.\d{3})+$"))
        .then(x.str.replace_all(r"\.", ""))
        .when(x.str.contains(r"^\d{1,3}(?:,\d{3})+$"))
        .then(x.str.replace_all(",", ""))
        .when(x.str.contains(r",") & ~x.str.contains(r"\."))
        .then(x.str.replace_all(",", "."))
        .otherwise(x)
        .cast(pl.Float64, strict=False)
    )


def construir_proyectos(datos, vigencia):
    """Extrae proyectos/gestiones desde la columna de texto del Plan Indicativo."""
    prog_ff = datos["prog_fisica_financiera"]

    # La vigencia 2026 tiene dos columnas candidatas; se prefiere la que contenga datos.
    col_proyecto = f"PROYECTOS {vigencia}"
    if col_proyecto not in prog_ff.columns:
        return pl.DataFrame()

    texto = pl.col(col_proyecto)

    patron_bpin = r"\((?i:bpin)\s*:\s*([^()]+?)\s*\)"
    patron_tipo_banco = r"\((?i:tipo\s+de\s+banco)\s*:\s*([^()]+?)\s*\)"
    patron_meta = (
        r"\((?i:(?:"
        r"meta\s+del\s+proyecto|"
        r"meta\s+de\s+la\s+gesti(?:ón|on)|"
        r"meta\s+total\s+del\s+indicador|"
        r"meta\s+total\s+de\s+la\s+vigencia|"
        r"meta\s+total\s+del\s+proyecto|"
        r"meta\s+programada|"
        r"meta\s+de\s+la\s+vigencia"
        r"))\s*:\s*([^()]+?)\s*\)"
    )
    # NOTA: el notebook usa literal "2024" en este patrón (línea 845 del notebook).
    # En la app generalizamos a \d{4} para que la extracción funcione también en
    # vigencia 2025/2026. Si el texto solo trae "Ejecución 2024", el comportamiento
    # es idéntico al del notebook.
    patron_ejecutado = (
        r"\((?i:(?:"
        r"ejecuci(?:ón|on)\s+\d{4}|"
        r"ejecutado|"
        r"ejecuci(?:ón|on)"
        r"))\s*:\s*([^()]+?)\s*\)"
    )
    patron_estado = r"\((?i:estado\s+en\s+portafolio)\s*:\s*([^()]+?)\s*\)"
    patron_bloques_info = (
        r"\((?i:(?:"
        r"bpin|"
        r"tipo\s+de\s+banco|"
        r"meta\s+del\s+proyecto|"
        r"meta\s+de\s+la\s+gesti(?:ón|on)|"
        r"meta\s+total\s+del\s+indicador|"
        r"meta\s+total\s+de\s+la\s+vigencia|"
        r"meta\s+total\s+del\s+proyecto|"
        r"meta\s+programada|"
        r"meta\s+de\s+la\s+vigencia|"
        r"ejecuci(?:ón|on)\s+\d{4}|"
        r"ejecutado|"
        r"ejecuci(?:ón|on)|"
        r"estado\s+en\s+portafolio"
        r"))\s*:\s*[^()]+?\s*\)"
    )

    return (
        prog_ff
        .select(
            "Codigo Meta", "Línea Estratégica", "Sector PDD", "Programa PDD",
            "Indicador de producto principal", "código de indicador principal",
            col_proyecto,
        )
        .with_columns(
            pl.col(col_proyecto).fill_null("").cast(pl.String).alias(col_proyecto)
        )
        .filter(pl.col(col_proyecto) != "", pl.col(col_proyecto) != "0")
        .with_columns(
            pl.col(col_proyecto)
            .str.replace_all(r"\n\s*\n+", "\n\n")
            .str.split("\n\n")
            .alias(col_proyecto)
        )
        .explode(col_proyecto)
        .with_columns(pl.col(col_proyecto).str.strip_chars().alias(col_proyecto))
        .filter(pl.col(col_proyecto) != "", pl.col(col_proyecto) != "0")
        .with_columns(
            texto.str.replace_all(patron_bloques_info, "")
                 .str.replace_all(r"\s+", " ")
                 .str.strip_chars()
                 .alias("Nombre del Proyecto"),
            _extraer_regex(texto, patron_bpin).alias("BPIN"),
            _extraer_regex(texto, patron_tipo_banco).alias("Tipo de Banco"),
            _normalizar_numero(_extraer_regex(texto, patron_meta)).alias("Meta"),
            _normalizar_numero(_extraer_regex(texto, patron_ejecutado)).alias("Ejecutado"),
            _extraer_regex(texto, patron_estado).alias("Estado en portafolio"),
        )
        .drop(col_proyecto)
    )


# =========================================================================
# Exportación a Excel — proyectos (vigencia única o consolidado)
# =========================================================================
def generar_excel_proyectos(
    df_proyectos: pd.DataFrame,
    titulo: str,
    subtitulo: str,
) -> bytes:
    """Genera un xlsx formateado corporativamente con la tabla de proyectos.

    df_proyectos debe traer las columnas que ya construimos en la pestaña +
    opcionalmente 'Vigencia PI' al inicio (cuando es consolidado de todas
    las vigencias).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    XL_BLUE_DARK = "003D6C"
    XL_ORANGE    = "CF7000"
    XL_INK       = "0D1B2A"
    XL_INK_MUTE  = "4A5A6A"
    XL_HAIRLINE  = "E2E0D8"
    XL_BEIGE     = "F0EEE9"
    XL_PAPER     = "FCFCFB"
    XL_ALT_ROW   = "F7F7F4"

    thin = Border(
        left=Side(style="thin", color=XL_HAIRLINE),
        right=Side(style="thin", color=XL_HAIRLINE),
        top=Side(style="thin", color=XL_HAIRLINE),
        bottom=Side(style="thin", color=XL_HAIRLINE),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    ws.sheet_view.showGridLines = False

    n_cols = len(df_proyectos.columns)

    # --- Masthead ---
    ws.cell(row=1, column=1, value=titulo)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1)
    c.font = Font(name="Montserrat", bold=True, size=18, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 38

    ws.cell(row=2, column=1, value=subtitulo)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1)
    c.font = Font(name="Open Sans", italic=True, size=10, color=XL_INK_MUTE)
    c.fill = PatternFill("solid", fgColor=XL_PAPER)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    # Línea naranja decorativa
    for col in range(1, n_cols + 1):
        c = ws.cell(row=3, column=col)
        c.fill = PatternFill("solid", fgColor=XL_ORANGE)
    ws.row_dimensions[3].height = 4

    # --- Encabezado tabla ---
    header_row = 5
    for c_idx, col_name in enumerate(df_proyectos.columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=str(col_name))
        cell.font = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=XL_BLUE_DARK),
            right=Side(style="thin", color=XL_BLUE_DARK),
            top=Side(style="thin", color=XL_BLUE_DARK),
            bottom=Side(style="medium", color=XL_ORANGE),
        )
    ws.row_dimensions[header_row].height = 32

    # --- Cuerpo ---
    pct_cols = {"Avance"}
    num_cols = {"Meta", "Ejecutado"}
    for r_idx, (_, row) in enumerate(df_proyectos.iterrows(), start=header_row + 1):
        alt = (r_idx - header_row) % 2 == 0
        for c_idx, col_name in enumerate(df_proyectos.columns, start=1):
            v = row[col_name]
            if pd.isna(v):
                v = None
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = Font(name="Open Sans", size=10, color=XL_INK)
            if alt:
                cell.fill = PatternFill("solid", fgColor=XL_ALT_ROW)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_name in pct_cols:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in num_cols:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # --- Anchos ---
    anchos = {
        "Vigencia PI": 12, "Codigo Meta": 14, "Línea Estratégica": 28,
        "Sector PDD": 24, "Programa PDD": 32,
        "Nombre del Proyecto": 50, "BPIN": 18, "Indicador": 42,
        "Tipo de Banco": 22, "Meta": 14, "Ejecutado": 14, "Avance": 12,
    }
    for c_idx, col_name in enumerate(df_proyectos.columns, start=1):
        letra = get_column_letter(c_idx)
        if col_name in anchos:
            ws.column_dimensions[letra].width = anchos[col_name]
        else:
            serie = df_proyectos[col_name].astype(str)
            max_len = max(len(str(col_name)), serie.str.len().max() if not serie.empty else 0)
            ws.column_dimensions[letra].width = min(max(12, max_len + 2), 45)

    # Congelar paneles bajo el encabezado
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def construir_dataframe_proyectos_listo(datos: dict, vigencia: str) -> pd.DataFrame:
    """Toma construir_proyectos y agrega columnas calculadas (Indicador, Avance)
    en el formato listo para mostrar/exportar."""
    df = construir_proyectos(datos, vigencia).to_pandas()
    if df.empty:
        return df

    df["Avance"] = df.apply(
        lambda r: (r["Ejecutado"] / r["Meta"])
        if pd.notna(r["Meta"]) and pd.notna(r["Ejecutado"]) and r["Meta"] != 0
        else None,
        axis=1,
    )

    def _fmt(row):
        codigo = row.get("código de indicador principal")
        nombre = row.get("Indicador de producto principal")
        codigo = "" if pd.isna(codigo) else str(codigo).strip()
        nombre = "" if pd.isna(nombre) else str(nombre).strip()
        if codigo and nombre:
            return f"{codigo} — {nombre}"
        return codigo or nombre or ""

    df["Indicador"] = df.apply(_fmt, axis=1)
    return df


# =========================================================================
# Exportación a Excel con formato corporativo
# =========================================================================
def generar_reporte_excel(
    datos: dict,
    vigencia: str,
    ejec_financ_tipo,
    ejec_acumulada_tipo,
    categorias_pdd,
    ejec_dependencia,
    avances_fisicos,
) -> bytes:
    """Genera un Excel con todas las hojas del reporte, formateado con la paleta corporativa."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ---- Paleta en formato openpyxl (sin '#') ----
    XL_BLUE_DARK  = "003D6C"
    XL_BLUE       = "1754AB"
    XL_ORANGE     = "CF7000"
    XL_ORANGE_LT  = "D88C16"
    XL_GREEN_LT   = "17743D"
    XL_PAPER      = "FBFAF6"
    XL_BEIGE      = "F1EDE2"
    XL_HAIRLINE   = "D9D4C7"
    XL_INK        = "0D1B2A"
    XL_INK_MUTE   = "4A5A6A"
    XL_BAR_BG     = "F1EDE2"

    # ---- Estilos base ----
    thin_border = Border(
        left=Side(style="thin", color=XL_HAIRLINE),
        right=Side(style="thin", color=XL_HAIRLINE),
        top=Side(style="thin", color=XL_HAIRLINE),
        bottom=Side(style="thin", color=XL_HAIRLINE),
    )

    def estilo_header(cell):
        cell.font = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color=XL_BLUE_DARK),
            right=Side(style="thin", color=XL_BLUE_DARK),
            top=Side(style="thin", color=XL_BLUE_DARK),
            bottom=Side(style="medium", color=XL_ORANGE),
        )

    def estilo_dato(cell, alt=False):
        cell.font = Font(name="Open Sans", size=10, color=XL_INK)
        if alt:
            cell.fill = PatternFill("solid", fgColor="FAF8F2")
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    def estilo_total(cell):
        cell.font = Font(name="Montserrat", bold=True, color=XL_BLUE_DARK, size=10)
        cell.fill = PatternFill("solid", fgColor=XL_BEIGE)
        cell.border = Border(
            left=Side(style="thin", color=XL_HAIRLINE),
            right=Side(style="thin", color=XL_HAIRLINE),
            top=Side(style="medium", color=XL_BLUE_DARK),
            bottom=Side(style="thin", color=XL_HAIRLINE),
        )
        cell.alignment = Alignment(vertical="center")

    def escribir_tabla(ws, df: pd.DataFrame, start_row: int, fila_total: dict = None,
                       columnas_pct=None, columnas_money=None, columnas_num=None):
        """Escribe un DataFrame con estilos institucionales desde start_row."""
        columnas_pct = columnas_pct or []
        columnas_money = columnas_money or []
        columnas_num = columnas_num or []

        # Encabezados
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=c_idx, value=str(col_name))
            estilo_header(cell)
        ws.row_dimensions[start_row].height = 32

        # Filas de datos
        for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
            alt = (r_idx - start_row) % 2 == 0
            for c_idx, col_name in enumerate(df.columns, start=1):
                v = row[col_name]
                if pd.isna(v):
                    v = None
                cell = ws.cell(row=r_idx, column=c_idx, value=v)
                estilo_dato(cell, alt=alt)
                if col_name in columnas_pct:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_money:
                    cell.number_format = '"$ "#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_num:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        last_row = start_row + len(df)

        # Fila de total
        if fila_total:
            last_row += 1
            for c_idx, col_name in enumerate(df.columns, start=1):
                v = fila_total.get(col_name, "")
                if c_idx == 1 and not v:
                    v = "TOTAL"
                cell = ws.cell(row=last_row, column=c_idx, value=v)
                estilo_total(cell)
                if col_name in columnas_pct:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_money:
                    cell.number_format = '"$ "#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in columnas_num:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        return last_row

    def ajustar_anchos(ws, df: pd.DataFrame, anchos_especificos=None):
        anchos_especificos = anchos_especificos or {}
        for c_idx, col_name in enumerate(df.columns, start=1):
            letra = get_column_letter(c_idx)
            if col_name in anchos_especificos:
                ws.column_dimensions[letra].width = anchos_especificos[col_name]
            else:
                serie = df[col_name].astype(str)
                max_len = max(len(str(col_name)), serie.str.len().max() if not serie.empty else 0)
                ws.column_dimensions[letra].width = min(max(12, max_len + 2), 45)

    def agregar_titulo(ws, titulo: str, subtitulo: str, start_row: int = 1, span: int = 6):
        # Franja superior azul oscura con el título
        ws.cell(row=start_row, column=1, value=titulo)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span)
        tcell = ws.cell(row=start_row, column=1)
        tcell.font = Font(name="Montserrat", bold=True, size=16, color="FFFFFF")
        tcell.fill = PatternFill("solid", fgColor=XL_BLUE_DARK)
        tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row].height = 36

        # Franja de eyebrow naranja
        ws.cell(row=start_row + 1, column=1, value=subtitulo)
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=span)
        scell = ws.cell(row=start_row + 1, column=1)
        scell.font = Font(name="Open Sans", italic=True, size=10, color=XL_INK_MUTE)
        scell.fill = PatternFill("solid", fgColor=XL_PAPER)
        scell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row + 1].height = 22

        # Franja naranja divisoria de 2 px equivalentes
        ws.cell(row=start_row + 2, column=1, value="")
        ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=span)
        dcell = ws.cell(row=start_row + 2, column=1)
        dcell.fill = PatternFill("solid", fgColor=XL_ORANGE)
        ws.row_dimensions[start_row + 2].height = 4

        return start_row + 4  # fila siguiente útil

    # ---- Crear workbook ----
    wb = Workbook()

    # =======================================================
    # Hoja 1: Portada
    # =======================================================
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False

    # Cabecera estilo masthead
    ws.cell(row=2, column=2, value="INFORME DE SEGUIMIENTO")
    ws.merge_cells("B2:G2")
    c = ws.cell(row=2, column=2)
    c.font = Font(name="Montserrat", bold=True, size=9, color=XL_ORANGE)
    c.alignment = Alignment(horizontal="left")

    ws.cell(row=4, column=2, value="Plan Indicativo 2024—2027")
    ws.merge_cells("B4:G4")
    c = ws.cell(row=4, column=2)
    c.font = Font(name="Montserrat", bold=True, size=28, color=XL_BLUE_DARK)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[4].height = 42

    ws.cell(row=5, column=2, value=f"Vigencia en análisis: {vigencia}")
    ws.merge_cells("B5:G5")
    c = ws.cell(row=5, column=2)
    c.font = Font(name="Open Sans", italic=True, size=12, color=XL_INK_MUTE)
    c.alignment = Alignment(horizontal="left")

    # Línea naranja decorativa
    for col in range(2, 8):
        c = ws.cell(row=7, column=col)
        c.fill = PatternFill("solid", fgColor=XL_ORANGE)
    ws.row_dimensions[7].height = 4

    # Bloque de descripción
    ws.cell(row=9, column=2,
            value="Consolida la programación y ejecución de los indicadores de producto "
                  "y sus fuentes de financiación. Los datos de 2024 y 2025 corresponden "
                  "a vigencias cerradas; los archivos de 2026 se actualizan en el "
                  "repositorio del sistema.")
    ws.merge_cells("B9:G11")
    c = ws.cell(row=9, column=2)
    c.font = Font(name="Open Sans", size=11, color=XL_INK)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Tarjeta KPI
    kpis = [
        (f"Programación {vigencia}",
         ejec_financ_tipo.select(pl.col(f"Programación Financiera {vigencia}").sum()).item() or 0,
         "money"),
        (f"Ejecución {vigencia}",
         ejec_financ_tipo.select(pl.col(f"Ejecución Financiera {vigencia}").sum()).item() or 0,
         "money"),
        ("Programación Cuatrienio",
         ejec_acumulada_tipo.select(pl.col("Programación Cuatrienio").sum()).item() or 0,
         "money"),
        ("Ejecución Acumulada",
         ejec_acumulada_tipo.select(pl.col("Ejecución Financiera Acumulada").sum()).item() or 0,
         "money"),
        (f"Avance ponderado {vigencia}",
         avances_fisicos["avance_vig_ponderado"] or 0, "pct"),
        ("Avance ponderado cuatrienio",
         avances_fisicos["avance_cuatrienio_total"] or 0, "pct"),
    ]

    fila_kpi = 14
    for i, (label, valor, tipo) in enumerate(kpis):
        col = 2 + (i % 3) * 2
        row = fila_kpi + (i // 3) * 4

        # Etiqueta
        ws.cell(row=row, column=col, value=label.upper())
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Montserrat", bold=True, size=8, color=XL_INK_MUTE)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(
            left=Side(style="medium", color=XL_BLUE),
            top=Side(style="thin", color=XL_HAIRLINE),
            right=Side(style="thin", color=XL_HAIRLINE),
        )

        # Valor
        ws.cell(row=row + 1, column=col, value=valor)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        c = ws.cell(row=row + 1, column=col)
        c.font = Font(name="Montserrat", bold=True, size=18, color=XL_INK)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(
            left=Side(style="medium", color=XL_BLUE),
            bottom=Side(style="thin", color=XL_HAIRLINE),
            right=Side(style="thin", color=XL_HAIRLINE),
        )
        if tipo == "money":
            c.number_format = '"$ "#,##0'
        elif tipo == "pct":
            c.number_format = "0.00%"

    # Anchos
    for col_letra in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letra].width = 18
    ws.column_dimensions["A"].width = 3

    # =======================================================
    # Hoja 2: Ejecución Financiera por Fuente — Vigencia
    # =======================================================
    ws = wb.create_sheet("Financiera por Fuente")
    ws.sheet_view.showGridLines = False

    df_fin_vig = ejec_financ_tipo.to_pandas()[
        ["Tipo Fuente", "Clasificación Recursos",
         f"Programación Financiera {vigencia}",
         f"Ejecución Financiera {vigencia}",
         "Porcentaje de Ejecución Financiera"]
    ].rename(columns={
        f"Programación Financiera {vigencia}": f"Programación {vigencia}",
        f"Ejecución Financiera {vigencia}": f"Ejecución {vigencia}",
        "Porcentaje de Ejecución Financiera": "% Ejecución",
    })

    inicio = agregar_titulo(
        ws,
        f"Ejecución financiera — Vigencia {vigencia}",
        "Programación vs ejecución por clasificación de recursos",
        span=len(df_fin_vig.columns),
    )

    total_prog = df_fin_vig[f"Programación {vigencia}"].sum()
    total_ejec = df_fin_vig[f"Ejecución {vigencia}"].sum()
    fila_tot = {
        f"Programación {vigencia}": total_prog,
        f"Ejecución {vigencia}": total_ejec,
        "% Ejecución": (total_ejec / total_prog) if total_prog else 0,
    }
    escribir_tabla(
        ws, df_fin_vig, inicio, fila_total=fila_tot,
        columnas_money=[f"Programación {vigencia}", f"Ejecución {vigencia}"],
        columnas_pct=["% Ejecución"],
    )
    ajustar_anchos(ws, df_fin_vig, anchos_especificos={
        "Tipo Fuente": 32, "Clasificación Recursos": 28,
        f"Programación {vigencia}": 22, f"Ejecución {vigencia}": 22, "% Ejecución": 16,
    })

    # =======================================================
    # Hoja 3: Ejecución Financiera Cuatrienio
    # =======================================================
    ws = wb.create_sheet("Financiera Cuatrienio")
    ws.sheet_view.showGridLines = False

    df_acum = ejec_acumulada_tipo.to_pandas().copy()
    df_acum["% Avance Cuatrienio"] = df_acum.apply(
        lambda r: (r["Ejecución Financiera Acumulada"] / r["Programación Cuatrienio"])
        if r["Programación Cuatrienio"] else 0, axis=1
    )
    df_acum = df_acum[[
        "Tipo Fuente", "Clasificación Recursos",
        "Programación Cuatrienio",
        "Ejecución Financiera 2024", "Ejecución Financiera 2025", "Ejecución Financiera 2026",
        "Ejecución Financiera Acumulada", "% Avance Cuatrienio",
    ]]

    inicio = agregar_titulo(
        ws, "Ejecución financiera — Cuatrienio",
        "Ejecución acumulada por fuente desde 2024 hasta 2026",
        span=len(df_acum.columns),
    )

    fila_tot = {
        "Programación Cuatrienio": df_acum["Programación Cuatrienio"].sum(),
        "Ejecución Financiera 2024": df_acum["Ejecución Financiera 2024"].sum(),
        "Ejecución Financiera 2025": df_acum["Ejecución Financiera 2025"].sum(),
        "Ejecución Financiera 2026": df_acum["Ejecución Financiera 2026"].sum(),
        "Ejecución Financiera Acumulada": df_acum["Ejecución Financiera Acumulada"].sum(),
        "% Avance Cuatrienio": (
            df_acum["Ejecución Financiera Acumulada"].sum() / df_acum["Programación Cuatrienio"].sum()
            if df_acum["Programación Cuatrienio"].sum() else 0
        ),
    }
    money_cols = ["Programación Cuatrienio", "Ejecución Financiera 2024",
                  "Ejecución Financiera 2025", "Ejecución Financiera 2026",
                  "Ejecución Financiera Acumulada"]
    escribir_tabla(ws, df_acum, inicio, fila_total=fila_tot,
                   columnas_money=money_cols, columnas_pct=["% Avance Cuatrienio"])
    ajustar_anchos(ws, df_acum, anchos_especificos={"Tipo Fuente": 32, "Clasificación Recursos": 28})

    # =======================================================
    # Hoja 4: Ejecución Financiera por Categorías del PDD
    # =======================================================
    ws = wb.create_sheet("Por Categoría PDD")
    ws.sheet_view.showGridLines = False

    fila = 1
    for cat_key, cat_label, col_grupo in [
        ("lineas",    "Líneas Estratégicas", "Línea Estratégica"),
        ("sectores",  "Sectores PDD",        "Sector PDD"),
        ("programas", "Programas PDD",       "Programa PDD"),
    ]:
        df_c = categorias_pdd[cat_key].to_pandas()[
            [col_grupo, f"Programación Financiera {vigencia}",
             f"Ejecución Financiera {vigencia}", "Porcentaje de Ejecución Financiera"]
        ].rename(columns={
            f"Programación Financiera {vigencia}": f"Programación {vigencia}",
            f"Ejecución Financiera {vigencia}": f"Ejecución {vigencia}",
            "Porcentaje de Ejecución Financiera": "% Ejecución",
        })
        fila = agregar_titulo(
            ws, cat_label, f"Programación vs ejecución financiera — {vigencia}",
            start_row=fila, span=len(df_c.columns),
        )
        tot_prog = df_c[f"Programación {vigencia}"].sum()
        tot_ejec = df_c[f"Ejecución {vigencia}"].sum()
        fila_tot = {
            f"Programación {vigencia}": tot_prog,
            f"Ejecución {vigencia}": tot_ejec,
            "% Ejecución": (tot_ejec / tot_prog) if tot_prog else 0,
        }
        ultima = escribir_tabla(
            ws, df_c, fila, fila_total=fila_tot,
            columnas_money=[f"Programación {vigencia}", f"Ejecución {vigencia}"],
            columnas_pct=["% Ejecución"],
        )
        fila = ultima + 3  # espacio entre tablas

    # Ajuste de anchos (columnas típicas)
    anchos = {"A": 48, "B": 22, "C": 22, "D": 16}
    for letra, w in anchos.items():
        ws.column_dimensions[letra].width = w

    # =======================================================
    # Hoja 5: Ejecución Física por Línea/Sector
    # =======================================================
    ws = wb.create_sheet("Ejecución Física")
    ws.sheet_view.showGridLines = False

    fila = 1
    for grupo, df_src, col_grupo, titulo in [
        ("Líneas Estratégicas — Vigencia",
         avances_fisicos["avance_vig_lineas"].to_pandas(), "Línea Estratégica",
         f"Avance físico por Línea Estratégica — {vigencia}"),
        ("Líneas Estratégicas — Cuatrienio",
         avances_fisicos["avance_cuatri_lineas"].to_pandas(), "Línea Estratégica",
         "Avance físico por Línea Estratégica — Cuatrienio"),
        ("Sectores PDD — Vigencia",
         avances_fisicos["avance_vig_sectores"].to_pandas(), "Sector PDD",
         f"Avance físico por Sector PDD — {vigencia}"),
        ("Sectores PDD — Cuatrienio",
         avances_fisicos["avance_cuatri_sectores"].to_pandas(), "Sector PDD",
         "Avance físico por Sector PDD — Cuatrienio"),
    ]:
        df_g = df_src[[col_grupo, "% Aporte Cumplimiento PDD",
                       "Sobre Numero Total de Indicadores", "% Eficacia Operativa"]].copy()
        df_g = df_g.rename(columns={
            "% Aporte Cumplimiento PDD": "Aporte PDD",
            "Sobre Numero Total de Indicadores": "Peso relativo",
            "% Eficacia Operativa": "Eficacia Operativa",
        })
        fila = agregar_titulo(ws, grupo, titulo, start_row=fila, span=len(df_g.columns))
        ultima = escribir_tabla(
            ws, df_g, fila,
            columnas_pct=["Aporte PDD", "Peso relativo", "Eficacia Operativa"],
        )
        fila = ultima + 3

    ws.column_dimensions["A"].width = 48
    for letra in ["B", "C", "D"]:
        ws.column_dimensions[letra].width = 20

    # =======================================================
    # Hoja 6: Ejecución por Dependencia
    # =======================================================
    ws = wb.create_sheet("Por Dependencia")
    ws.sheet_view.showGridLines = False

    df_dep = ejec_dependencia.to_pandas()[
        ["Varias Secretarías", "Dependencia Responsable",
         f"Metas Programadas {vigencia}", f"Metas Cumplidas al 100% {vigencia}",
         f"Porcentaje de Ejecución {vigencia}", "Porcentaje de Ejecución Acumulada"]
    ].rename(columns={
        f"Metas Programadas {vigencia}": f"Programadas {vigencia}",
        f"Metas Cumplidas al 100% {vigencia}": "Cumplidas 100%",
        f"Porcentaje de Ejecución {vigencia}": f"Avance {vigencia}",
        "Porcentaje de Ejecución Acumulada": "Avance acumulado",
    })

    inicio = agregar_titulo(
        ws, "Ejecución por Dependencia Responsable",
        f"Desempeño físico por secretaría — {vigencia}",
        span=len(df_dep.columns),
    )
    escribir_tabla(
        ws, df_dep, inicio,
        columnas_pct=[f"Avance {vigencia}", "Avance acumulado"],
    )
    ajustar_anchos(ws, df_dep, anchos_especificos={
        "Varias Secretarías": 22, "Dependencia Responsable": 40,
    })

    # =======================================================
    # Hoja 7: Detalle por Meta
    # =======================================================
    ws = wb.create_sheet("Detalle por Meta")
    ws.sheet_view.showGridLines = False

    df_det = datos["prog_fisica_financiera"].to_pandas()[[
        "Codigo Meta", "Línea Estratégica", "Sector PDD", "Programa PDD",
        "Indicador de producto principal", "Responsable",
        "Meta de cuatrenio",
        f"Meta Física Esperada {vigencia}",
        f"EJECUCIÓN {vigencia}",
        f"PORCENTAJE DE EJECUCIÓN {vigencia}",
        "EJECUCIÓN ACUMULADA", "PORCENTAJE DE EJECUCIÓN ACUMULADA",
        "CATEGORÍA DE EJECUCIÓN ACUMULADA",
    ]].rename(columns={
        "Indicador de producto principal": "Indicador",
        f"Meta Física Esperada {vigencia}": f"Meta {vigencia}",
        f"EJECUCIÓN {vigencia}": f"Ejecución {vigencia}",
        f"PORCENTAJE DE EJECUCIÓN {vigencia}": f"Avance {vigencia}",
        "EJECUCIÓN ACUMULADA": "Ejec. acumulada",
        "PORCENTAJE DE EJECUCIÓN ACUMULADA": "Avance acumulado",
        "CATEGORÍA DE EJECUCIÓN ACUMULADA": "Categoría",
    })

    inicio = agregar_titulo(
        ws, "Detalle por Meta",
        f"Inventario de indicadores de producto — Vigencia {vigencia}",
        span=len(df_det.columns),
    )
    escribir_tabla(
        ws, df_det, inicio,
        columnas_pct=[f"Avance {vigencia}", "Avance acumulado"],
        columnas_num=["Meta de cuatrenio", f"Meta {vigencia}",
                      f"Ejecución {vigencia}", "Ejec. acumulada"],
    )
    ws.freeze_panes = ws.cell(row=inicio + 1, column=1)
    ajustar_anchos(ws, df_det, anchos_especificos={
        "Codigo Meta": 14, "Línea Estratégica": 30, "Sector PDD": 26,
        "Programa PDD": 36, "Indicador": 40, "Responsable": 28, "Categoría": 18,
    })

    # =======================================================
    # Hoja 8: Proyectos
    # =======================================================
    df_proy = construir_proyectos(datos, vigencia).to_pandas()
    if not df_proy.empty:
        ws = wb.create_sheet("Proyectos")
        ws.sheet_view.showGridLines = False

        # Construye Indicador como "código — nombre" (igual que en la pestaña)
        def _fmt_ind(row):
            cod = row.get("código de indicador principal")
            nom = row.get("Indicador de producto principal")
            cod = "" if pd.isna(cod) else str(cod).strip()
            nom = "" if pd.isna(nom) else str(nom).strip()
            if cod and nom:
                return f"{cod} — {nom}"
            return cod or nom or ""

        df_proy_xl = df_proy.copy()
        df_proy_xl["Indicador"] = df_proy_xl.apply(_fmt_ind, axis=1)
        df_proy_xl = df_proy_xl[[
            "Codigo Meta", "Línea Estratégica", "Sector PDD", "Programa PDD",
            "Nombre del Proyecto", "BPIN", "Indicador",
            "Tipo de Banco", "Meta", "Ejecutado",
        ]]
        df_proy_xl["Avance"] = df_proy_xl.apply(
            lambda r: (r["Ejecutado"] / r["Meta"])
            if pd.notna(r["Meta"]) and pd.notna(r["Ejecutado"]) and r["Meta"] != 0
            else None,
            axis=1,
        )

        inicio = agregar_titulo(
            ws, "Proyectos y Gestiones",
            f"Inventario extraído del Plan Indicativo — Vigencia {vigencia}",
            span=len(df_proy_xl.columns),
        )
        escribir_tabla(
            ws, df_proy_xl, inicio,
            columnas_pct=["Avance"],
            columnas_num=["Meta", "Ejecutado"],
        )
        ws.freeze_panes = ws.cell(row=inicio + 1, column=1)
        ajustar_anchos(ws, df_proy_xl, anchos_especificos={
            "Codigo Meta": 14, "Nombre del Proyecto": 50, "Indicador": 42,
            "Tipo de Banco": 22,
            "Línea Estratégica": 28, "Sector PDD": 24, "Programa PDD": 32,
        })

    # ---- Serialización ----
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# =========================================================================
# Sidebar
# =========================================================================
st.sidebar.markdown(
    f"""
    <div style='padding: 0.8rem 0 1.2rem 0; border-bottom: 1px solid rgba(255,255,255,0.1); margin-bottom: 1rem;'>
        <div style='font-family: {FONT_MONO}, monospace; font-size: 0.68rem; letter-spacing: 0.22em;
                    color: {COLORS["orange"]}; text-transform: uppercase;'>
            Plan de Desarrollo
        </div>
        <div style='font-family: {FONT_DISPLAY}, {FONT_HEADING}, sans-serif; font-size: 2rem;
                    font-weight: 400; color: #fff; line-height: 1; margin-top: 0.25rem;'>
            2024<span style='color: {COLORS["orange"]}'>—</span>2027
        </div>
        <div style='font-family: {FONT_HEADING}, sans-serif; font-size: 0.7rem;
                    color: #b9c6d6; margin-top: 0.5rem; letter-spacing: 0.14em;
                    text-transform: uppercase; font-weight: 500;'>
            Sistema de Seguimiento
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.sidebar.markdown("### Archivos actualizables")
st.sidebar.markdown(
    f"""
    <div style='font-family: {FONT_BODY}, sans-serif; font-size: 0.78rem;
                color: #b9c6d6; margin: -0.4rem 0 0.8rem 0; line-height: 1.5;'>
        Las vigencias 2024 y 2025 ya cerraron y se consultan del repositorio.
        Sube aquí los archivos de 2026 y el Plan Indicativo.
    </div>
    """,
    unsafe_allow_html=True,
)

# Botón global de recarga: siempre visible, independiente del modo de carga.
# Limpia toda la caché (descargas + procesamiento) para forzar lectura fresca.
if st.sidebar.button("Recargar datos del repositorio", use_container_width=True,
                     help="Limpia la caché y vuelve a descargar los archivos desde GitHub"):
    st.cache_data.clear()
    # Limpia también el archivo Excel pre-generado, por si la vigencia cambia
    st.session_state.pop("xlsx_bytes", None)
    st.session_state.pop("xlsx_vigencia", None)
    st.rerun()

modo_carga = st.sidebar.radio(
    "Modo",
    options=["Usar datos del repositorio", "Subir archivos 2026 y Plan Indicativo"],
    index=0,
    label_visibility="collapsed",
)

archivos_bytes = {}

# Vigencias cerradas: siempre del repo
try:
    with st.spinner("Cargando vigencias cerradas (2024-2025)..."):
        for key in ARCHIVOS_CERRADOS:
            archivos_bytes[key] = descargar_desde_github(GH[key])
except Exception as e:
    st.sidebar.error(f"Error al cargar vigencias cerradas: {e}")
    st.stop()

# Archivos actualizables
if modo_carga == "Usar datos del repositorio":
    try:
        with st.spinner("Descargando Plan Indicativo y archivos 2026..."):
            for key in ARCHIVOS_ACTUALIZABLES:
                archivos_bytes[key] = descargar_desde_github(GH[key])
    except Exception as e:
        st.sidebar.error(f"Error al descargar: {e}")
        st.stop()
else:
    st.sidebar.markdown("---")
    st.sidebar.markdown("**Plan Indicativo**")
    pi_file = st.sidebar.file_uploader(
        "Plan Indicativo 2024-2027",
        type=["xlsx"], key="pi_upload",
        label_visibility="collapsed",
    )

    st.sidebar.markdown("**Hacienda 2026**")
    h26_file = st.sidebar.file_uploader(
        "Ejecución Hacienda 2026",
        type=["xlsx"], key="h26_upload",
        label_visibility="collapsed",
    )

    st.sidebar.markdown("**Regalías 2026**")
    r26_file = st.sidebar.file_uploader(
        "Pagos Regalías 2026",
        type=["xlsx"], key="r26_upload",
        label_visibility="collapsed",
    )

    if not (pi_file and h26_file and r26_file):
        st.warning(
            "Sube los tres archivos requeridos: Plan Indicativo, Hacienda 2026 y Regalías 2026."
        )
        st.stop()

    archivos_bytes["pi"]  = pi_file.getvalue()
    archivos_bytes["h26"] = h26_file.getvalue()
    archivos_bytes["r26"] = r26_file.getvalue()

# Procesamiento
try:
    datos = procesar_datos(
        archivos_bytes["pi"], archivos_bytes["h24"], archivos_bytes["r24"],
        archivos_bytes["h25"], archivos_bytes["r25"],
        archivos_bytes["ads_rp_25"], archivos_bytes["ads_reg_25"],
        archivos_bytes["gestiones_25"], archivos_bytes["fondo_mixto_25"],
        archivos_bytes["inder_25"],
        archivos_bytes["h26"], archivos_bytes["r26"],
    )
except Exception as e:
    st.error(f"Error procesando los archivos: {e}")
    st.exception(e)
    st.stop()

st.sidebar.markdown("### Vigencia de análisis")
filtro_vigencia = st.sidebar.selectbox(
    "Vigencia",
    options=["2024", "2025", "2026"],
    index=2,
    label_visibility="collapsed",
)

# =========================================================================
# Masthead editorial
# =========================================================================
st.markdown(
    f"""
    <div class="masthead">
        <div>
            <div class="eyebrow">Informe de Seguimiento  /  Número {filtro_vigencia[-2:]}</div>
            <h1>Plan <em>Indicativo</em></h1>
        </div>
        <div class="edition">
            <strong>Vigencia en análisis</strong><br/>
            {filtro_vigencia} · Cuatrienio 2024—2027<br/>
            Ejecución física y financiera
        </div>
    </div>
    <div class="subhead">
        Instrumento de seguimiento al cumplimiento de metas del Plan de Desarrollo.
        Consolida la programación y ejecución de los indicadores de producto y sus fuentes de financiación.
    </div>
    """,
    unsafe_allow_html=True,
)

# =========================================================================
# Cálculos
# =========================================================================
ejec_financ_tipo = construir_ejecucion_financ_tipo(datos, filtro_vigencia)
ejec_acumulada_tipo = construir_ejecucion_acumulada_tipo(datos)
categorias_pdd = construir_prog_financ_categorias(datos, filtro_vigencia)
ejec_dependencia = construir_ejec_por_dependencia(datos, filtro_vigencia)
avances_fisicos = construir_avances_fisicos(datos, filtro_vigencia)

# =========================================================================
# KPIs
# =========================================================================
prog_vig = ejec_financ_tipo.select(pl.col(f"Programación Financiera {filtro_vigencia}").sum()).item() or 0
ejec_vig = ejec_financ_tipo.select(pl.col(f"Ejecución Financiera {filtro_vigencia}").sum()).item() or 0
pct_vig = (ejec_vig / prog_vig) if prog_vig else 0

prog_cuatri = ejec_acumulada_tipo.select(pl.col("Programación Cuatrienio").sum()).item() or 0
ejec_acum = ejec_acumulada_tipo.select(pl.col("Ejecución Financiera Acumulada").sum()).item() or 0
pct_cuatri = (ejec_acum / prog_cuatri) if prog_cuatri else 0

c1, c2, c3, c4 = st.columns(4)
c1.metric(f"Programación {filtro_vigencia}", formato_pesos(prog_vig),
          help=TOOLTIPS["prog_vigencia"])
c2.metric(f"Ejecución {filtro_vigencia}", formato_pesos(ejec_vig), formato_porcentaje(pct_vig),
          help=TOOLTIPS["ejec_vigencia"])
c3.metric("Programación Cuatrienio", formato_pesos(prog_cuatri),
          help=TOOLTIPS["prog_cuatrienio"])
c4.metric("Ejecución Acumulada", formato_pesos(ejec_acum), formato_porcentaje(pct_cuatri),
          help=TOOLTIPS["ejec_acumulada"])

st.markdown("<hr/>", unsafe_allow_html=True)

# =========================================================================
# Pestañas
# =========================================================================
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "Ejecución Física",
    "Ejecución Financiera",
    "Distribución de Metas",
    "Ejecución por Dependencia",
    "Proyectos",
    "Exportar",
])

# -----------------------------------------------------------------
# 01. EJECUCIÓN FÍSICA
# -----------------------------------------------------------------
with tab1:
    seccion("01", "Ejecución Física",
            "Avance ponderado del cumplimiento de metas físicas del Plan de Desarrollo.",
            tooltip=(
                "Mide qué tanto se han cumplido las metas del Plan en términos "
                "físicos, no monetarios. El avance global se construye combinando "
                "el desempeño de cada programa con su peso dentro del Plan: los "
                "programas con más metas pesan más en el resultado. La 'Eficacia "
                "Operativa' permite comparar líneas y sectores ajustando por su "
                "tamaño relativo, de modo que dependencias pequeñas con buen "
                "desempeño no quedan invisibilizadas frente a las más grandes."
            ))

    k1, k2 = st.columns(2)
    k1.metric(f"Avance ponderado — Vigencia {filtro_vigencia}",
              formato_porcentaje(avances_fisicos["avance_vig_ponderado"] or 0),
              help=TOOLTIPS["avance_vig_ponderado"])
    k2.metric("Avance ponderado — Cuatrienio",
              formato_porcentaje(avances_fisicos["avance_cuatrienio_total"] or 0),
              help=TOOLTIPS["avance_cuatrienio_ponderado"])

    st.markdown(" ")
    sub_v, sub_c = st.tabs([f"Vigencia {filtro_vigencia}", "Cuatrienio"])

    def fig_bar_horizontal(df, cat_col, val_col, titulo, color_scale):
        df2 = df.sort_values(val_col, ascending=True)
        fig = px.bar(
            df2, x=val_col, y=cat_col,
            orientation="h", text=val_col,
            color=val_col, color_continuous_scale=color_scale,
            title=titulo,
        )
        fig.update_traces(texttemplate="%{text:.1%}", textposition="outside",
                          marker_line_color=COLORS["blue_dark"], marker_line_width=0.5)
        fig.update_layout(xaxis_tickformat=".0%",
                          height=max(450, len(df2) * 32),
                          showlegend=False, coloraxis_showscale=False, bargap=0.3)
        return fig

    columnas_lineas = [
        {"key": "Línea Estratégica", "label": "Línea Estratégica", "type": "text"},
        {"key": "% Aporte Cumplimiento PDD", "label": "Aporte PDD", "type": "pct"},
        {"key": "Sobre Numero Total de Indicadores", "label": "Peso relativo", "type": "pct"},
        {"key": "% Eficacia Operativa", "label": "Eficacia Operativa", "type": "pctbar"},
    ]
    columnas_sectores = [
        {"key": "Sector PDD", "label": "Sector PDD", "type": "text"},
        {"key": "% Aporte Cumplimiento PDD", "label": "Aporte PDD", "type": "pct"},
        {"key": "Sobre Numero Total de Indicadores", "label": "Peso relativo", "type": "pct"},
        {"key": "% Eficacia Operativa", "label": "Eficacia Operativa", "type": "pctbar"},
    ]

    with sub_v:
        st.markdown("##### Por Línea Estratégica")
        vista = selector_vista("vista_fis_vig_lineas")
        df = avances_fisicos["avance_vig_lineas"].to_pandas()
        if not df.empty:
            render_vista(
                vista,
                fig_factory=lambda: fig_bar_horizontal(
                    df, "Línea Estratégica", "% Eficacia Operativa",
                    f"Eficacia Operativa por Línea Estratégica — {filtro_vigencia}", SCALE_BLUE),
                df_tabla=df,
                columnas=columnas_lineas,
            )
        else:
            st.info("Sin datos para la vigencia seleccionada.")

        st.markdown("##### Por Sector PDD")
        vista = selector_vista("vista_fis_vig_sectores")
        df = avances_fisicos["avance_vig_sectores"].to_pandas()
        if not df.empty:
            render_vista(
                vista,
                fig_factory=lambda: fig_bar_horizontal(
                    df, "Sector PDD", "% Eficacia Operativa",
                    f"Eficacia Operativa por Sector PDD — {filtro_vigencia}", SCALE_BLUE),
                df_tabla=df,
                columnas=columnas_sectores,
            )

    with sub_c:
        st.markdown("##### Por Línea Estratégica")
        vista = selector_vista("vista_fis_cuatri_lineas")
        df = avances_fisicos["avance_cuatri_lineas"].to_pandas()
        render_vista(
            vista,
            fig_factory=lambda: fig_bar_horizontal(
                df, "Línea Estratégica", "% Eficacia Operativa",
                "Eficacia Operativa por Línea Estratégica — Cuatrienio", SCALE_GREEN),
            df_tabla=df,
            columnas=columnas_lineas,
        )

        st.markdown("##### Por Sector PDD")
        vista = selector_vista("vista_fis_cuatri_sectores")
        df = avances_fisicos["avance_cuatri_sectores"].to_pandas()
        render_vista(
            vista,
            fig_factory=lambda: fig_bar_horizontal(
                df, "Sector PDD", "% Eficacia Operativa",
                "Eficacia Operativa por Sector PDD — Cuatrienio", SCALE_GREEN),
            df_tabla=df,
            columnas=columnas_sectores,
        )

# -----------------------------------------------------------------
# 02. EJECUCIÓN FINANCIERA
# -----------------------------------------------------------------
with tab2:
    seccion("02", "Ejecución Financiera",
            "Comportamiento de recursos programados frente a ejecutados por fuente y categoría del PDD.",
            tooltip=(
                "Compara los recursos presupuestados frente a los efectivamente "
                "pagados. La programación reúne las diez fuentes de financiación "
                "del Plan (recursos propios, SGP, Regalías, cofinanciaciones, "
                "crédito y otras). La ejecución consolida los reportes de "
                "Hacienda, Regalías y, para 2025, también los de Aguas de Sucre, "
                "Gestiones, PDET, Fondo Mixto e Indersucre. El % de ejecución "
                "indica qué tanto se ha utilizado de cada fuente."
            ))

    sub_v, sub_c = st.tabs([f"Vigencia {filtro_vigencia}", "Cuatrienio"])

    with sub_v:
        k1, k2, k3 = st.columns(3)
        k1.metric("Programación", formato_pesos(prog_vig),
                  help=TOOLTIPS["prog_vigencia"])
        k2.metric("Ejecución", formato_pesos(ejec_vig),
                  help=TOOLTIPS["ejec_vigencia"])
        k3.metric("Avance", formato_porcentaje(pct_vig),
                  help=TOOLTIPS["avance_vigencia"])

        # --- Por clasificación de recursos ---
        st.markdown("##### Por Clasificación de Recursos")
        vista = selector_vista("vista_fin_vig_tipo")
        df_tipo = ejec_financ_tipo.to_pandas()

        def fig_tipo():
            fig = go.Figure()
            fig.add_trace(go.Bar(
                name="Programación", x=df_tipo["Clasificación Recursos"],
                y=df_tipo[f"Programación Financiera {filtro_vigencia}"],
                marker=dict(color=COLORS["blue_dark"], line=dict(color="#fff", width=1)),
            ))
            fig.add_trace(go.Bar(
                name="Ejecución", x=df_tipo["Clasificación Recursos"],
                y=df_tipo[f"Ejecución Financiera {filtro_vigencia}"],
                marker=dict(color=COLORS["orange_deep"], line=dict(color="#fff", width=1)),
            ))
            fig.update_layout(
                barmode="group", height=460,
                title=f"Programación vs Ejecución por Fuente — {filtro_vigencia}",
                yaxis_title="Valor (COP)", xaxis_tickangle=-25, bargap=0.25,
            )
            return fig

        columnas_tipo = [
            {"key": "Clasificación Recursos", "label": "Fuente", "type": "text"},
            {"key": "Tipo Fuente", "label": "Tipo", "type": "text"},
            {"key": f"Programación Financiera {filtro_vigencia}", "label": f"Programación {filtro_vigencia}", "type": "money"},
            {"key": f"Ejecución Financiera {filtro_vigencia}", "label": f"Ejecución {filtro_vigencia}", "type": "money"},
            {"key": "Porcentaje de Ejecución Financiera", "label": "Avance", "type": "pctbar"},
        ]
        totales_tipo = {
            f"Programación Financiera {filtro_vigencia}": df_tipo[f"Programación Financiera {filtro_vigencia}"].sum(),
            f"Ejecución Financiera {filtro_vigencia}": df_tipo[f"Ejecución Financiera {filtro_vigencia}"].sum(),
            "Porcentaje de Ejecución Financiera": (
                df_tipo[f"Ejecución Financiera {filtro_vigencia}"].sum()
                / df_tipo[f"Programación Financiera {filtro_vigencia}"].sum()
                if df_tipo[f"Programación Financiera {filtro_vigencia}"].sum() else 0
            ),
        }
        render_vista(vista, fig_factory=fig_tipo, df_tabla=df_tipo,
                     columnas=columnas_tipo, totales=totales_tipo)

        # --- Por categorías del PDD ---
        st.markdown("##### Por Categorías del Plan de Desarrollo")
        cat1, cat2, cat3 = st.tabs(["Líneas Estratégicas", "Sectores PDD", "Programas PDD"])

        def fig_cat(df, col_cat, titulo):
            fig = go.Figure()
            fig.add_trace(go.Bar(
                name="Programación", x=df[col_cat],
                y=df[f"Programación Financiera {filtro_vigencia}"],
                marker_color=COLORS["blue_dark"],
            ))
            fig.add_trace(go.Bar(
                name="Ejecución", x=df[col_cat],
                y=df[f"Ejecución Financiera {filtro_vigencia}"],
                marker_color=COLORS["orange_deep"],
            ))
            fig.update_layout(barmode="group", height=480,
                              title=titulo, xaxis_tickangle=-30)
            return fig

        with cat1:
            vista = selector_vista("vista_fin_cat_lineas")
            df = categorias_pdd["lineas"].to_pandas()
            columnas = [
                {"key": "Línea Estratégica", "label": "Línea Estratégica", "type": "text"},
                {"key": f"Programación Financiera {filtro_vigencia}", "label": "Programación", "type": "money"},
                {"key": f"Ejecución Financiera {filtro_vigencia}", "label": "Ejecución", "type": "money"},
                {"key": "Porcentaje de Ejecución Financiera", "label": "Avance", "type": "pctbar"},
            ]
            totales = {
                f"Programación Financiera {filtro_vigencia}": df[f"Programación Financiera {filtro_vigencia}"].sum(),
                f"Ejecución Financiera {filtro_vigencia}": df[f"Ejecución Financiera {filtro_vigencia}"].sum(),
                "Porcentaje de Ejecución Financiera": (
                    df[f"Ejecución Financiera {filtro_vigencia}"].sum()
                    / df[f"Programación Financiera {filtro_vigencia}"].sum()
                    if df[f"Programación Financiera {filtro_vigencia}"].sum() else 0
                ),
            }
            render_vista(
                vista,
                fig_factory=lambda: fig_cat(df, "Línea Estratégica",
                                            f"Programación vs Ejecución por Línea — {filtro_vigencia}"),
                df_tabla=df, columnas=columnas, totales=totales,
            )

        with cat2:
            vista = selector_vista("vista_fin_cat_sectores")
            df = categorias_pdd["sectores"].to_pandas()
            columnas = [
                {"key": "Sector PDD", "label": "Sector PDD", "type": "text"},
                {"key": f"Programación Financiera {filtro_vigencia}", "label": "Programación", "type": "money"},
                {"key": f"Ejecución Financiera {filtro_vigencia}", "label": "Ejecución", "type": "money"},
                {"key": "Porcentaje de Ejecución Financiera", "label": "Avance", "type": "pctbar"},
            ]
            totales = {
                f"Programación Financiera {filtro_vigencia}": df[f"Programación Financiera {filtro_vigencia}"].sum(),
                f"Ejecución Financiera {filtro_vigencia}": df[f"Ejecución Financiera {filtro_vigencia}"].sum(),
                "Porcentaje de Ejecución Financiera": (
                    df[f"Ejecución Financiera {filtro_vigencia}"].sum()
                    / df[f"Programación Financiera {filtro_vigencia}"].sum()
                    if df[f"Programación Financiera {filtro_vigencia}"].sum() else 0
                ),
            }
            render_vista(
                vista,
                fig_factory=lambda: fig_cat(df, "Sector PDD",
                                            f"Programación vs Ejecución por Sector — {filtro_vigencia}"),
                df_tabla=df, columnas=columnas, totales=totales,
            )

        with cat3:
            vista = selector_vista("vista_fin_cat_programas")
            df = categorias_pdd["programas"].to_pandas()
            columnas = [
                {"key": "Programa PDD", "label": "Programa PDD", "type": "text"},
                {"key": f"Programación Financiera {filtro_vigencia}", "label": "Programación", "type": "money"},
                {"key": f"Ejecución Financiera {filtro_vigencia}", "label": "Ejecución", "type": "money"},
                {"key": "Porcentaje de Ejecución Financiera", "label": "Avance", "type": "pctbar"},
            ]
            totales = {
                f"Programación Financiera {filtro_vigencia}": df[f"Programación Financiera {filtro_vigencia}"].sum(),
                f"Ejecución Financiera {filtro_vigencia}": df[f"Ejecución Financiera {filtro_vigencia}"].sum(),
                "Porcentaje de Ejecución Financiera": (
                    df[f"Ejecución Financiera {filtro_vigencia}"].sum()
                    / df[f"Programación Financiera {filtro_vigencia}"].sum()
                    if df[f"Programación Financiera {filtro_vigencia}"].sum() else 0
                ),
            }

            def fig_programas():
                df_top = df.sort_values(f"Ejecución Financiera {filtro_vigencia}", ascending=True).tail(20)
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    name="Programación", y=df_top["Programa PDD"],
                    x=df_top[f"Programación Financiera {filtro_vigencia}"],
                    orientation="h",
                    marker_color=COLORS["blue_dark"],
                ))
                fig.add_trace(go.Bar(
                    name="Ejecución", y=df_top["Programa PDD"],
                    x=df_top[f"Ejecución Financiera {filtro_vigencia}"],
                    orientation="h",
                    marker_color=COLORS["orange_deep"],
                ))
                fig.update_layout(
                    barmode="group", height=650,
                    title=f"Top 20 programas por ejecución — {filtro_vigencia}",
                    xaxis_title="Valor (COP)",
                )
                return fig

            render_vista(vista, fig_factory=fig_programas,
                         df_tabla=df, columnas=columnas, totales=totales)

    with sub_c:
        k1, k2, k3 = st.columns(3)
        k1.metric("Programación Cuatrienio", formato_pesos(prog_cuatri),
                  help=TOOLTIPS["prog_cuatrienio"])
        k2.metric("Ejecución Acumulada", formato_pesos(ejec_acum),
                  help=TOOLTIPS["ejec_acumulada"])
        k3.metric("Avance", formato_porcentaje(pct_cuatri),
                  help=TOOLTIPS["avance_cuatrienio"])

        vista = selector_vista("vista_fin_cuatri")
        df_acum = ejec_acumulada_tipo.to_pandas()

        def fig_acum():
            fig = go.Figure()
            fig.add_trace(go.Bar(name="Ejecución 2024", x=df_acum["Clasificación Recursos"],
                                 y=df_acum["Ejecución Financiera 2024"], marker_color=COLORS["green_light"]))
            fig.add_trace(go.Bar(name="Ejecución 2025", x=df_acum["Clasificación Recursos"],
                                 y=df_acum["Ejecución Financiera 2025"], marker_color=COLORS["blue"]))
            fig.add_trace(go.Bar(name="Ejecución 2026", x=df_acum["Clasificación Recursos"],
                                 y=df_acum["Ejecución Financiera 2026"], marker_color=COLORS["orange_deep"]))
            fig.update_layout(barmode="stack", height=520,
                              title="Ejecución Acumulada por Fuente (2024-2026)",
                              yaxis_title="Valor (COP)", xaxis_tickangle=-25, bargap=0.25)
            return fig

        # Calcular % de avance cuatrienio por fuente
        df_acum_tabla = df_acum.copy()
        df_acum_tabla["% Avance Cuatrienio"] = df_acum_tabla.apply(
            lambda r: (r["Ejecución Financiera Acumulada"] / r["Programación Cuatrienio"])
            if r["Programación Cuatrienio"] else 0, axis=1
        )

        columnas_acum = [
            {"key": "Clasificación Recursos", "label": "Fuente", "type": "text"},
            {"key": "Programación Cuatrienio", "label": "Programación Cuatrienio", "type": "money"},
            {"key": "Ejecución Financiera 2024", "label": "Ejec. 2024", "type": "money"},
            {"key": "Ejecución Financiera 2025", "label": "Ejec. 2025", "type": "money"},
            {"key": "Ejecución Financiera 2026", "label": "Ejec. 2026", "type": "money"},
            {"key": "Ejecución Financiera Acumulada", "label": "Ejec. Acumulada", "type": "money"},
            {"key": "% Avance Cuatrienio", "label": "Avance", "type": "pctbar"},
        ]
        totales_acum = {
            "Programación Cuatrienio": df_acum_tabla["Programación Cuatrienio"].sum(),
            "Ejecución Financiera 2024": df_acum_tabla["Ejecución Financiera 2024"].sum(),
            "Ejecución Financiera 2025": df_acum_tabla["Ejecución Financiera 2025"].sum(),
            "Ejecución Financiera 2026": df_acum_tabla["Ejecución Financiera 2026"].sum(),
            "Ejecución Financiera Acumulada": df_acum_tabla["Ejecución Financiera Acumulada"].sum(),
            "% Avance Cuatrienio": (
                df_acum_tabla["Ejecución Financiera Acumulada"].sum()
                / df_acum_tabla["Programación Cuatrienio"].sum()
                if df_acum_tabla["Programación Cuatrienio"].sum() else 0
            ),
        }
        render_vista(vista, fig_factory=fig_acum, df_tabla=df_acum_tabla,
                     columnas=columnas_acum, totales=totales_acum)

# -----------------------------------------------------------------
# 03. DISTRIBUCIÓN DE METAS
# -----------------------------------------------------------------
with tab3:
    seccion("03", "Distribución de Metas",
            "Peso relativo de la programación física en cada vigencia del cuatrienio.",
            tooltip=(
                "Muestra cómo se reparte el cumplimiento físico del Plan entre "
                "los cuatro años: cuánto se planea cumplir cada vigencia frente "
                "a la meta total del cuatrienio. La suma de los porcentajes "
                "puede no dar exactamente 100% porque algunas metas son "
                "acumulativas y otras corresponden a flujos que se reinician "
                "cada año."
            ))

    prog_ff = datos["prog_fisica_financiera"]
    programacion_cuatrienio = prog_ff.select(pl.col("Meta de cuatrenio").sum()).item() or 1

    distribucion = {}
    for v in ["2024", "2025", "2026", "2027"]:
        suma = prog_ff.select(pl.col(f"Meta Física Esperada {v}").sum()).item() or 0
        distribucion[v] = (suma / programacion_cuatrienio, suma)

    df_dist = pd.DataFrame({
        "Vigencia": list(distribucion.keys()),
        "Suma metas físicas": [v[1] for v in distribucion.values()],
        "Distribución": [v[0] for v in distribucion.values()],
    })

    vista = selector_vista("vista_distr")

    def fig_distr():
        fig = go.Figure(data=[go.Pie(
            labels=df_dist["Vigencia"], values=df_dist["Distribución"],
            hole=0.55,
            marker=dict(
                colors=[COLORS["blue_dark"], COLORS["cyan"],
                        COLORS["orange_deep"], COLORS["brown"]],
                line=dict(color="#fff", width=2),
            ),
            textinfo="label+percent",
            textfont=dict(family=FONT_HEADING, size=14, color="#fff"),
        )])
        fig.update_layout(height=470, showlegend=False,
                          title="Distribución por Vigencia")
        return fig

    columnas_dist = [
        {"key": "Vigencia", "label": "Vigencia", "type": "text"},
        {"key": "Suma metas físicas", "label": "Metas físicas programadas", "type": "int"},
        {"key": "Distribución", "label": "Distribución", "type": "pctbar"},
    ]
    totales_dist = {
        "Suma metas físicas": df_dist["Suma metas físicas"].sum(),
        "Distribución": df_dist["Distribución"].sum(),
    }
    render_vista(vista, fig_factory=fig_distr,
                 df_tabla=df_dist, columnas=columnas_dist, totales=totales_dist)

    st.markdown(" ")
    st.markdown("##### Conteo de Metas")
    a, b = st.columns(2)
    a.metric("Total de indicadores de producto",
             formato_entero(avances_fisicos['numero_total_metas']))
    b.metric(f"Indicadores con programación en {filtro_vigencia}",
             formato_entero(avances_fisicos['numero_metas_prog_vigencia']))

# -----------------------------------------------------------------
# 04. EJECUCIÓN POR DEPENDENCIA
# -----------------------------------------------------------------
with tab4:
    seccion("04", "Ejecución por Dependencia",
            "Desempeño de las dependencias responsables de la ejecución del Plan de Desarrollo.",
            tooltip=(
                "Para cada secretaría o dependencia se reporta cuántas metas "
                "tiene programadas en la vigencia, cuántas alcanzaron el 100% "
                "(categoría 'Superior'), su avance promedio en la vigencia y su "
                "avance promedio acumulado del cuatrienio. Las dependencias se "
                "homologan según la tabla oficial del Plan Indicativo, que "
                "permite agrupar variantes de nombre y, cuando aplica, marcar "
                "responsabilidades compartidas entre varias secretarías."
            ))

    df_dep = ejec_dependencia.to_pandas()

    varias_opciones = sorted([x for x in df_dep["Varias Secretarías"].dropna().unique()])
    if varias_opciones:
        filtro_sec = st.multiselect(
            "Filtrar por agrupación (Varias Secretarías)",
            options=varias_opciones, default=[],
        )
        if filtro_sec:
            df_dep = df_dep[df_dep["Varias Secretarías"].isin(filtro_sec)]

    if df_dep.empty:
        st.info("No hay dependencias con metas programadas en la vigencia seleccionada.")
    else:
        vista = selector_vista("vista_dep")

        def fig_dep():
            df_plot = df_dep.sort_values(f"Porcentaje de Ejecución {filtro_vigencia}", ascending=True)
            fig = px.bar(
                df_plot, x=f"Porcentaje de Ejecución {filtro_vigencia}", y="Dependencia Responsable",
                orientation="h", text=f"Porcentaje de Ejecución {filtro_vigencia}",
                color=f"Porcentaje de Ejecución {filtro_vigencia}", color_continuous_scale=SCALE_ORANGE,
                title=f"Porcentaje de Ejecución Física por Dependencia — {filtro_vigencia}",
            )
            fig.update_traces(texttemplate="%{text:.1%}", textposition="outside",
                              marker_line_color=COLORS["brown"], marker_line_width=0.5)
            fig.update_layout(xaxis_tickformat=".0%",
                              height=max(450, len(df_plot) * 32),
                              showlegend=False, coloraxis_showscale=False, bargap=0.25)
            return fig

        columnas_dep = [
            {"key": "Dependencia Responsable", "label": "Dependencia", "type": "text"},
            {"key": f"Metas Programadas {filtro_vigencia}", "label": f"Programadas {filtro_vigencia}", "type": "int"},
            {"key": f"Metas Cumplidas al 100% {filtro_vigencia}", "label": "Cumplidas 100%", "type": "int"},
            {"key": f"Porcentaje de Ejecución {filtro_vigencia}", "label": f"Avance {filtro_vigencia}", "type": "pctbar"},
            {"key": "Porcentaje de Ejecución Acumulada", "label": "Avance acumulado", "type": "pctbar"},
        ]
        render_vista(vista, fig_factory=fig_dep, df_tabla=df_dep, columnas=columnas_dep)

# -----------------------------------------------------------------
# 05. PROYECTOS
# -----------------------------------------------------------------
with tab5:
    seccion("05", "Proyectos",
            "Inventario de proyectos y gestiones asociadas a las metas del Plan de Desarrollo, "
            "extraídos de la columna de texto del Plan Indicativo.",
            tooltip=(
                "Lista los proyectos y gestiones registrados para la vigencia "
                "en el Plan Indicativo. De cada uno se extraen el código BPIN, "
                "el indicador de producto al que aporta, el tipo de banco al "
                "que pertenece (Banco de Proyectos, Banco de Programas, etc.), "
                "la meta física comprometida y lo ejecutado. El avance se "
                "reporta en unidades físicas (no en pesos): qué tanto del "
                "producto o servicio comprometido se entregó."
            ))

    df_proy = construir_dataframe_proyectos_listo(datos, filtro_vigencia)

    # ---- Botones de descarga (siempre visibles arriba) ----
    st.markdown("##### Descargar inventario")
    dl1, dl2, _ = st.columns([1, 1, 2])

    with dl1:
        if not df_proy.empty:
            cols_export = [
                "Codigo Meta", "Línea Estratégica", "Sector PDD", "Programa PDD",
                "Nombre del Proyecto", "BPIN", "Indicador", "Tipo de Banco",
                "Meta", "Ejecutado", "Avance",
            ]
            df_export_vig = df_proy.reindex(columns=[c for c in cols_export if c in df_proy.columns])
            xlsx_vig = generar_excel_proyectos(
                df_export_vig,
                titulo=f"Proyectos y Gestiones — Vigencia {filtro_vigencia}",
                subtitulo=f"Inventario completo extraído del Plan Indicativo",
            )
            st.download_button(
                f"Descargar vigencia {filtro_vigencia}",
                data=xlsx_vig,
                file_name=f"proyectos_{filtro_vigencia}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_proy_vig",
                use_container_width=True,
            )
        else:
            st.button(f"Descargar vigencia {filtro_vigencia}", disabled=True,
                      use_container_width=True, key="dl_proy_vig_disabled")

    with dl2:
        # Consolidado de las cuatro vigencias del Plan
        try:
            partes = []
            for v in ["2024", "2025", "2026", "2027"]:
                df_v = construir_dataframe_proyectos_listo(datos, v)
                if not df_v.empty:
                    df_v = df_v.copy()
                    df_v.insert(0, "Vigencia PI", v)
                    partes.append(df_v)
            if partes:
                df_consol = pd.concat(partes, ignore_index=True)
                cols_consol = [
                    "Vigencia PI", "Codigo Meta", "Línea Estratégica", "Sector PDD",
                    "Programa PDD", "Nombre del Proyecto", "BPIN", "Indicador",
                    "Tipo de Banco", "Meta", "Ejecutado", "Avance",
                ]
                df_consol = df_consol.reindex(columns=[c for c in cols_consol if c in df_consol.columns])
                xlsx_all = generar_excel_proyectos(
                    df_consol,
                    titulo="Proyectos y Gestiones — Consolidado del Cuatrienio",
                    subtitulo="Inventario unificado de todas las vigencias del Plan (2024–2027)",
                )
                st.download_button(
                    "Descargar todas las vigencias",
                    data=xlsx_all,
                    file_name="proyectos_2024-2027.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_proy_all",
                    use_container_width=True,
                )
            else:
                st.button("Descargar todas las vigencias", disabled=True,
                          use_container_width=True, key="dl_proy_all_disabled")
        except Exception as e:
            st.error(f"No se pudo generar el consolidado: {e}")

    st.markdown("<hr/>", unsafe_allow_html=True)

    if df_proy.empty:
        st.info(f"No hay proyectos ni gestiones registrados para la vigencia {filtro_vigencia}.")
    else:
        # ---- Conteo por tipo de banco (sustituye a las tarjetas monetarias) ----
        total_registros = len(df_proy)

        conteo_bancos = (
            df_proy["Tipo de Banco"].fillna("Sin clasificar")
            .replace("", "Sin clasificar")
            .value_counts()
        )
        # Tomamos los tres tipos de banco con más registros + total general
        tipos_top = conteo_bancos.head(3)

        columnas_kpi = st.columns(1 + len(tipos_top))
        columnas_kpi[0].metric("Total proyectos/gestiones", formato_entero(total_registros),
                               help=TOOLTIPS["total_proyectos_gestiones"])
        for i, (banco, conteo) in enumerate(tipos_top.items(), start=1):
            columnas_kpi[i].metric(banco, formato_entero(conteo))

        st.markdown(" ")

        # ---- Filtros ----
        fp1, fp2, fp3 = st.columns(3)
        with fp1:
            lineas_p = ["(Todas)"] + sorted(df_proy["Línea Estratégica"].dropna().unique().tolist())
            sel_linea_p = st.selectbox("Línea Estratégica", lineas_p, key="proy_linea")
        with fp2:
            df_tp = df_proy if sel_linea_p == "(Todas)" else df_proy[df_proy["Línea Estratégica"] == sel_linea_p]
            sectores_p = ["(Todos)"] + sorted(df_tp["Sector PDD"].dropna().unique().tolist())
            sel_sector_p = st.selectbox("Sector PDD", sectores_p, key="proy_sector")
        with fp3:
            df_tp2 = df_tp if sel_sector_p == "(Todos)" else df_tp[df_tp["Sector PDD"] == sel_sector_p]
            bancos = ["(Todos)"] + sorted([b for b in df_tp2["Tipo de Banco"].dropna().unique().tolist() if b])
            sel_banco = st.selectbox("Tipo de Banco", bancos, key="proy_banco")

        df_proy_f = df_proy.copy()
        if sel_linea_p != "(Todas)":
            df_proy_f = df_proy_f[df_proy_f["Línea Estratégica"] == sel_linea_p]
        if sel_sector_p != "(Todos)":
            df_proy_f = df_proy_f[df_proy_f["Sector PDD"] == sel_sector_p]
        if sel_banco != "(Todos)":
            df_proy_f = df_proy_f[df_proy_f["Tipo de Banco"] == sel_banco]

        st.caption(f"Mostrando {len(df_proy_f):,} proyectos/gestiones")

        columnas_proy = [
            {"key": "Codigo Meta", "label": "Meta", "type": "text"},
            {"key": "Nombre del Proyecto", "label": "Proyecto / Gestión", "type": "text"},
            {"key": "BPIN", "label": "BPIN", "type": "text"},
            {"key": "Indicador", "label": "Indicador de Producto", "type": "text"},
            {"key": "Tipo de Banco", "label": "Banco", "type": "text"},
            {"key": "Meta", "label": "Meta física", "type": "num2"},
            {"key": "Ejecutado", "label": "Ejecutado", "type": "num2"},
            {"key": "Avance", "label": "Avance", "type": "pctbar"},
        ]
        render_table(df_proy_f.head(200), columnas_proy)

        if len(df_proy_f) > 200:
            st.caption("La tabla muestra los primeros 200 registros. Usa los botones de descarga para el listado completo.")

# -----------------------------------------------------------------
# -----------------------------------------------------------------
# 06. EXPORTAR
# -----------------------------------------------------------------
with tab6:
    seccion("06", "Exportar",
            "Descarga un archivo Excel consolidado con toda la información del tablero, "
            "formateado con la identidad visual corporativa.")

    # Tarjeta descriptiva de lo que incluye el archivo
    st.markdown(
        f"""
        <div style='background:#fff; border:1px solid var(--hairline);
                    border-left: 3px solid {COLORS["orange_deep"]};
                    padding: 1.2rem 1.4rem; border-radius: 2px; margin-bottom: 1.2rem;'>
            <div style='font-family: {FONT_HEADING}, sans-serif; font-size: 0.72rem;
                        text-transform: uppercase; letter-spacing: 0.14em;
                        color: {COLORS["orange_deep"]}; font-weight: 700; margin-bottom: 0.5rem;'>
                Contenido del archivo
            </div>
            <div style='font-family: {FONT_BODY}, sans-serif; font-size: 0.9rem;
                        color: var(--ink); line-height: 1.7;'>
                El archivo incluye ocho hojas: <strong>Portada</strong> con los indicadores
                clave de la vigencia, <strong>Financiera por Fuente</strong>,
                <strong>Financiera Cuatrienio</strong>, <strong>Por Categoría PDD</strong>
                (líneas, sectores y programas), <strong>Ejecución Física</strong> con
                los avances por línea y sector, <strong>Por Dependencia</strong>,
                <strong>Detalle por Meta</strong> y <strong>Proyectos</strong>.
                Todas las tablas se entregan con la paleta corporativa, tipografías
                Montserrat y Open Sans, y formatos numéricos listos para imprimir.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_gen, col_info = st.columns([1, 2])
    with col_gen:
        st.markdown(
            f"""
            <div style='font-family: {FONT_HEADING}, sans-serif; font-size: 0.72rem;
                        text-transform: uppercase; letter-spacing: 0.14em;
                        color: var(--ink-mute); font-weight: 600; margin-bottom: 0.6rem;'>
                Vigencia a exportar
            </div>
            <div style='font-family: {FONT_DISPLAY}, {FONT_HEADING}, sans-serif;
                        font-size: 2.4rem; color: {COLORS["blue_dark"]}; font-weight: 700;
                        line-height: 1;'>
                {filtro_vigencia}
            </div>
            <div style='font-family: {FONT_MONO}, monospace; font-size: 0.72rem;
                        color: var(--ink-mute); margin-top: 0.4rem;'>
                Cambia la vigencia desde la barra lateral.
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_info:
        # Botón de generación y descarga
        if st.button("Generar archivo Excel", use_container_width=False, key="gen_xlsx"):
            with st.spinner("Generando archivo Excel con formato corporativo..."):
                try:
                    xlsx_bytes = generar_reporte_excel(
                        datos, filtro_vigencia,
                        ejec_financ_tipo, ejec_acumulada_tipo, categorias_pdd,
                        ejec_dependencia, avances_fisicos,
                    )
                    st.session_state["xlsx_bytes"] = xlsx_bytes
                    st.session_state["xlsx_vigencia"] = filtro_vigencia
                    st.success("Archivo generado correctamente. Usa el botón de descarga.")
                except Exception as e:
                    st.error(f"Error al generar el archivo: {e}")
                    st.exception(e)

        if (
            st.session_state.get("xlsx_bytes")
            and st.session_state.get("xlsx_vigencia") == filtro_vigencia
        ):
            st.download_button(
                "Descargar archivo Excel",
                data=st.session_state["xlsx_bytes"],
                file_name=f"Plan_Indicativo_{filtro_vigencia}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False,
                key="dl_xlsx",
            )


# =========================================================================
# Pie de página
# =========================================================================
st.markdown("<hr/>", unsafe_allow_html=True)
st.markdown(
    f"""
    <div style='display: flex; justify-content: space-between; align-items: center;
                padding: 0.6rem 0 1.2rem 0; font-size: 0.75rem; color: {COLORS["blue_dark"]};
                font-family: {FONT_MONO}, monospace; letter-spacing: 0.08em;
                border-top: 1px solid #d9d4c7;'>
        <span>Plan Indicativo · Sistema de Seguimiento 2024—2027</span>
        <span>Construido con Streamlit · Polars · Plotly</span>
    </div>
    """,
    unsafe_allow_html=True,
)
